# -*- coding: utf-8 -*-
"""
差旅费用详情记录 —— 本地桌面版
技术栈: Python 3 + tkinter + SQLite + Pillow + openpyxl
功能: 行程记录(日期/星期/出发地/到达地/交通方式/费用)、当天用餐与公共交通标记及截图、按日期筛选、导出 Excel
"""

import datetime
import os
import re
import shutil
import sqlite3
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import fitz  # PyMuPDF
from PIL import Image, ImageTk
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------- 路径常量 ----------------
if getattr(sys, 'frozen', False):
    # 打包成 exe 后:数据目录固定在 exe 所在位置,避免写入临时解压目录
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
SCREENSHOTS_DIR = os.path.join(DATA_DIR, 'screenshots')
EXPORT_DIR = os.path.join(BASE_DIR, '导出文件')
DB_PATH = os.path.join(DATA_DIR, '差旅记录.db')

TRANSPORT_OPTIONS = ['公共交通', '打车', '高铁', '飞机', '公共交通+打车', '甲方提供', '我方租车']
WEEKDAY_NAMES = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
IMG_EXTS = ('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.webp')
INVOICE_EXTS = IMG_EXTS + ('.pdf',)
FONT = ('Microsoft YaHei UI', 10)
FONT_BOLD = ('Microsoft YaHei UI', 10, 'bold')

# ---------------- 城市分类与费用标准 ----------------
# 一类城市
CITY_TIER1 = ['北京', '上海', '广州', '深圳']
# 二类城市: 省会 + 直辖市 + 计划单列市 + 经济发达城市
CITY_TIER2 = [
    # 省会 / 直辖市
    '天津', '重庆',
    '石家庄', '太原', '呼和浩特', '沈阳', '长春', '哈尔滨',
    '南京', '杭州', '合肥', '福州', '南昌', '济南', '郑州',
    '武汉', '长沙', '南宁', '海口', '成都', '贵阳', '昆明',
    '拉萨', '西安', '兰州', '西宁', '银川', '乌鲁木齐',
    # 计划单列市
    '大连', '青岛', '宁波', '厦门',
    # 经济发达城市
    '苏州', '无锡', '常州', '东莞', '佛山', '珠海', '温州',
    '烟台', '泉州', '潍坊',
]
# 三类城市: 其他地级市、县级市、县城

# 费用标准 (餐费, 市内交通)
EXPENSE_TIER1 = (80, 40)
EXPENSE_TIER2 = (60, 30)
EXPENSE_TIER3 = (50, 20)


def classify_city(city_name):
    """根据城市名判断类别, 返回 (tier_num, meal_rate, transit_rate)"""
    if not city_name:
        return (3, *EXPENSE_TIER3)
    city = city_name.strip().rstrip('市省区县')
    # 一类
    for c in CITY_TIER1:
        if city == c or c in city_name:
            return (1, *EXPENSE_TIER1)
    # 二类
    for c in CITY_TIER2:
        if city == c or c in city_name:
            return (2, *EXPENSE_TIER2)
    # 三类
    return (3, *EXPENSE_TIER3)


def parse_city_from_tag(tag):
    """从标签名解析城市名, 支持格式: 城市_日期 或 出差_城市_日期"""
    if not tag:
        return ''
    parts = tag.split('_')
    if not parts:
        return ''
    # 第一段如果不是纯中文标签(如"出差"),则视为城市名
    first = parts[0].strip()
    if first in ('出差', '差旅'):
        # 旧格式: 出差_城市_日期
        return parts[1] if len(parts) >= 2 else ''
    # 新格式: 城市_日期
    return first


# ---------------- 工具函数 ----------------

def parse_date(text):
    """解析 YYYY-MM-DD,非法返回 None"""
    try:
        return datetime.date.fromisoformat(text.strip())
    except (ValueError, AttributeError):
        return None


def weekday_of(date_str):
    """根据日期计算星期几"""
    d = parse_date(date_str)
    if d is None:
        return ''
    return WEEKDAY_NAMES[d.weekday()]


def rel_to_abs(rel_path):
    """数据库中的相对路径(正斜杠)转为绝对路径"""
    if not rel_path:
        return ''
    return os.path.join(BASE_DIR, *rel_path.split('/'))


# ---------------- 数据库层 ----------------

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def db_init():
    """初始化数据目录与数据库表"""
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(SCREENSHOTS_DIR, exist_ok=True)
    os.makedirs(EXPORT_DIR, exist_ok=True)
    conn = get_conn()
    try:
        conn.execute(
            """CREATE TABLE IF NOT EXISTS trips (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trip_date TEXT NOT NULL,
                weekday TEXT DEFAULT '', 
                depart TEXT DEFAULT '',
                arrive TEXT DEFAULT '',
                transport TEXT DEFAULT '',
                cost REAL DEFAULT 0,
                meal_flag INTEGER DEFAULT 0,
                meal_screenshot TEXT DEFAULT '',
                public_flag INTEGER DEFAULT 0,
                public_screenshot TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now', 'localtime'))
            )"""
        )
        conn.execute(
            """CREATE TABLE IF NOT EXISTS lodging (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                checkin_date TEXT NOT NULL,
                checkout_date TEXT NOT NULL,
                hotel TEXT DEFAULT '',
                amount REAL DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now', 'localtime'))
            )"""
        )
        conn.execute(
            """CREATE TABLE IF NOT EXISTS meals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                meal_date TEXT NOT NULL,
                meal_type TEXT DEFAULT '',
                place TEXT DEFAULT '',
                amount REAL DEFAULT 0,
                screenshot TEXT DEFAULT '',
                src_trip_id INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now', 'localtime'))
            )"""
        )
        conn.execute(
            """CREATE TABLE IF NOT EXISTS transports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                t_date TEXT NOT NULL,
                mode TEXT DEFAULT '',
                depart TEXT DEFAULT '',
                arrive TEXT DEFAULT '',
                amount REAL DEFAULT 0,
                public_flag INTEGER DEFAULT 0,
                screenshot TEXT DEFAULT '',
                src_trip_id INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now', 'localtime'))
            )"""
        )
        # 旧数据一次性迁移: trips 里的餐饮/交通字段迁到新表(幂等)
        migrate_legacy_trips(conn)
        # 为 trips 表增加 invoice 列(幂等, 已存在则忽略)
        try:
            conn.execute("ALTER TABLE trips ADD COLUMN invoice TEXT DEFAULT ''")
        except Exception:
            pass
        # 为四张表增加出差标签列(幂等)
        for _tbl in ('trips', 'lodging', 'meals', 'transports'):
            try:
                conn.execute("ALTER TABLE %s ADD COLUMN trip_tag TEXT DEFAULT ''" % _tbl)
            except Exception:
                pass
        # 附件表
        conn.execute(
            """CREATE TABLE IF NOT EXISTS trip_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trip_id INTEGER NOT NULL,
                file_type TEXT NOT NULL DEFAULT 'invoice',
                file_path TEXT NOT NULL,
                file_name TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now', 'localtime'))
            )"""
        )
        # 将 trips.invoice 旧字段数据迁移到 trip_files(幂等)
        rows = conn.execute("SELECT id, invoice FROM trips WHERE invoice IS NOT NULL AND invoice != ''").fetchall()
        for r in rows:
            exists = conn.execute(
                "SELECT 1 FROM trip_files WHERE trip_id=? AND file_path=?", (r['id'], r['invoice'])
            ).fetchone()
            if not exists:
                old_name = os.path.basename(r['invoice'])
                conn.execute(
                    "INSERT INTO trip_files (trip_id, file_type, file_path, file_name) VALUES (?,?,?,?)",
                    (r['id'], 'invoice', r['invoice'], old_name)
                )
        # 住宿附件表
        conn.execute(
            """CREATE TABLE IF NOT EXISTS lodging_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                lodging_id INTEGER NOT NULL,
                file_type TEXT NOT NULL DEFAULT 'invoice',
                file_path TEXT NOT NULL,
                file_name TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now', 'localtime'))
            )"""
        )
        conn.commit()
    finally:
        conn.close()


def add_trip(data):
    conn = get_conn()
    try:
        cur = conn.execute(
            """INSERT INTO trips (trip_date, weekday, depart, arrive, transport, cost,
                meal_flag, meal_screenshot, public_flag, public_screenshot, trip_tag)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (data['trip_date'], data['weekday'], data['depart'], data['arrive'],
             data['transport'], data['cost'], data['meal_flag'],
             data['meal_screenshot'], data['public_flag'], data['public_screenshot'],
             data.get('trip_tag', '')),
        )
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def update_trip(tid, data):
    conn = get_conn()
    try:
        conn.execute(
            """UPDATE trips SET trip_date=?, weekday=?, depart=?, arrive=?, transport=?,
                cost=?, meal_flag=?, meal_screenshot=?, public_flag=?, public_screenshot=?, trip_tag=?
               WHERE id=?""",
            (data['trip_date'], data['weekday'], data['depart'], data['arrive'],
             data['transport'], data['cost'], data['meal_flag'],
             data['meal_screenshot'], data['public_flag'], data['public_screenshot'],
             data.get('trip_tag', ''), tid),
        )
        conn.commit()
    finally:
        conn.close()


def get_trip(tid):
    conn = get_conn()
    try:
        return conn.execute('SELECT * FROM trips WHERE id=?', (tid,)).fetchone()
    finally:
        conn.close()


def delete_trip(tid, delete_files):
    """删除记录;delete_files 为真时同时删除关联截图和附件文件"""
    rec = get_trip(tid)
    if rec is None:
        return
    if delete_files:
        for rel in (rec['meal_screenshot'], rec['public_screenshot']):
            path = rel_to_abs(rel)
            if path and os.path.isfile(path):
                try:
                    os.remove(path)
                except OSError:
                    pass
        # 删除附件文件
        for f in get_trip_files(tid):
            remove_screenshot_file(f['file_path'])
    conn = get_conn()
    try:
        conn.execute('DELETE FROM trip_files WHERE trip_id=?', (tid,))
        conn.execute('DELETE FROM trips WHERE id=?', (tid,))
        conn.commit()
    finally:
        conn.close()


def query_trips(date_from='', date_to='', keyword='', tag=''):
    """按日期区间与关键字查询,日期倒序"""
    sql = 'SELECT * FROM trips WHERE 1=1'
    args = []
    if date_from:
        sql += ' AND trip_date >= ?'
        args.append(date_from)
    if date_to:
        sql += ' AND trip_date <= ?'
        args.append(date_to)
    if keyword:
        kw = '%' + keyword.strip().replace('%', '%%').replace('_', '\\_') + '%'
        sql += " AND (depart LIKE ? ESCAPE '\\' OR arrive LIKE ? ESCAPE '\\' OR transport LIKE ? ESCAPE '\\')"
        args.extend([kw, kw, kw])
    if tag:
        sql += ' AND trip_tag = ?'
        args.append(tag)
    sql += ' ORDER BY trip_date DESC, id DESC'
    conn = get_conn()
    try:
        return conn.execute(sql, args).fetchall()
    finally:
        conn.close()


# ---------- 附件管理(trip_files) ----------

def add_trip_file(trip_id, file_type, file_path, file_name=''):
    conn = get_conn()
    try:
        cur = conn.execute(
            'INSERT INTO trip_files (trip_id, file_type, file_path, file_name) VALUES (?,?,?,?)',
            (trip_id, file_type, file_path, file_name),
        )
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def get_trip_files(trip_id):
    conn = get_conn()
    try:
        return conn.execute(
            'SELECT * FROM trip_files WHERE trip_id=? ORDER BY id', (trip_id,)
        ).fetchall()
    finally:
        conn.close()


def delete_trip_file(fid, delete_file=True):
    if delete_file:
        conn = get_conn()
        try:
            row = conn.execute(
                'SELECT file_path FROM trip_files WHERE id=?', (fid,)
            ).fetchone()
            if row:
                remove_screenshot_file(row['file_path'])
        finally:
            conn.close()
    conn = get_conn()
    try:
        conn.execute('DELETE FROM trip_files WHERE id=?', (fid,))
        conn.commit()
    finally:
        conn.close()


def get_trip_file_counts(trip_ids):
    """批量获取行程的附件数量,返回 {trip_id: {'invoice': N, 'itinerary': N}}"""
    if not trip_ids:
        return {}
    placeholders = ','.join('?' * len(trip_ids))
    conn = get_conn()
    try:
        rows = conn.execute(
            'SELECT trip_id, file_type, COUNT(*) as cnt FROM trip_files '
            'WHERE trip_id IN (%s) GROUP BY trip_id, file_type' % placeholders,
            trip_ids,
        ).fetchall()
    finally:
        conn.close()
    result = {}
    for r in rows:
        tid = r['trip_id']
        if tid not in result:
            result[tid] = {'invoice': 0, 'itinerary': 0}
        result[tid][r['file_type']] = r['cnt']
    return result


# ---------- 住宿附件管理(lodging_files) ----------

def add_lodging_file(lodging_id, file_type, file_path, file_name=''):
    conn = get_conn()
    try:
        cur = conn.execute(
            'INSERT INTO lodging_files (lodging_id, file_type, file_path, file_name) VALUES (?,?,?,?)',
            (lodging_id, file_type, file_path, file_name),
        )
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def get_lodging_files(lodging_id):
    conn = get_conn()
    try:
        return conn.execute(
            'SELECT * FROM lodging_files WHERE lodging_id=? ORDER BY id', (lodging_id,)
        ).fetchall()
    finally:
        conn.close()


def delete_lodging_file(fid, delete_file=True):
    if delete_file:
        conn = get_conn()
        try:
            row = conn.execute(
                'SELECT file_path FROM lodging_files WHERE id=?', (fid,)
            ).fetchone()
            if row:
                remove_screenshot_file(row['file_path'])
        finally:
            conn.close()
    conn = get_conn()
    try:
        conn.execute('DELETE FROM lodging_files WHERE id=?', (fid,))
        conn.commit()
    finally:
        conn.close()


def get_lodging_file_counts(lodging_ids):
    """批量获取住宿的附件数量,返回 {lodging_id: {'invoice': N, 'receipt': N}}"""
    if not lodging_ids:
        return {}
    placeholders = ','.join('?' * len(lodging_ids))
    conn = get_conn()
    try:
        rows = conn.execute(
            'SELECT lodging_id, file_type, COUNT(*) as cnt FROM lodging_files '
            'WHERE lodging_id IN (%s) GROUP BY lodging_id, file_type' % placeholders,
            lodging_ids,
        ).fetchall()
    finally:
        conn.close()
    result = {}
    for r in rows:
        lid = r['lodging_id']
        if lid not in result:
            result[lid] = {'invoice': 0, 'receipt': 0}
        result[lid][r['file_type']] = r['cnt']
    return result


# ---------------- 住宿层 ----------------

def add_lodging(data):
    conn = get_conn()
    try:
        cur = conn.execute(
            """INSERT INTO lodging (checkin_date, checkout_date, hotel, amount, trip_tag)
               VALUES (?, ?, ?, ?, ?)""",
            (data['checkin_date'], data['checkout_date'], data['hotel'], data['amount'],
             data.get('trip_tag', '')),
        )
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def update_lodging(lid, data):
    conn = get_conn()
    try:
        conn.execute(
            """UPDATE lodging SET checkin_date=?, checkout_date=?, hotel=?, amount=?, trip_tag=?
               WHERE id=?""",
            (data['checkin_date'], data['checkout_date'], data['hotel'], data['amount'],
             data.get('trip_tag', ''), lid),
        )
        conn.commit()
    finally:
        conn.close()


def get_lodging(lid):
    conn = get_conn()
    try:
        return conn.execute('SELECT * FROM lodging WHERE id=?', (lid,)).fetchone()
    finally:
        conn.close()


def delete_lodging(lid):
    """删除住宿记录及其附件文件"""
    # 删除附件文件
    for f in get_lodging_files(lid):
        remove_screenshot_file(f['file_path'])
    conn = get_conn()
    try:
        conn.execute('DELETE FROM lodging_files WHERE lodging_id=?', (lid,))
        conn.execute('DELETE FROM lodging WHERE id=?', (lid,))
        conn.commit()
    finally:
        conn.close()


def query_lodging(date_from='', date_to='', keyword='', tag=''):
    """住宿记录,按入住日期区间与酒店关键字筛选,入住日期倒序"""
    sql = 'SELECT * FROM lodging WHERE 1=1'
    args = []
    if date_from:
        sql += ' AND checkin_date >= ?'
        args.append(date_from)
    if date_to:
        sql += ' AND checkin_date <= ?'
        args.append(date_to)
    if keyword:
        kw = '%' + keyword.strip().replace('%', '%%').replace('_', '\\_') + '%'
        sql += " AND (hotel LIKE ? ESCAPE '\\')"
        args.append(kw)
    if tag:
        sql += ' AND trip_tag = ?'
        args.append(tag)
    sql += ' ORDER BY checkin_date DESC, id DESC'
    conn = get_conn()
    try:
        return conn.execute(sql, args).fetchall()
    finally:
        conn.close()


def get_lodging_history():
    """最近用过的酒店名称,用于下拉记忆"""
    conn = get_conn()
    try:
        hotels = [r[0] for r in conn.execute(
            "SELECT hotel FROM lodging WHERE hotel!='' GROUP BY hotel ORDER BY MAX(id) DESC LIMIT 50")]
        return hotels
    finally:
        conn.close()


# ---------------- 餐饮 / 交通层 ----------------

MEAL_TYPES = ['早餐', '午餐', '晚餐', '下午茶', '夜宵', '其他']


def add_meal(data):
    conn = get_conn()
    try:
        cur = conn.execute(
            """INSERT INTO meals (meal_date, meal_type, place, amount, screenshot, trip_tag)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (data['meal_date'], data['meal_type'], data['place'], data['amount'],
             data.get('screenshot', ''), data.get('trip_tag', '')),
        )
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def update_meal(mid, data):
    conn = get_conn()
    try:
        conn.execute(
            """UPDATE meals SET meal_date=?, meal_type=?, place=?, amount=?, screenshot=?, trip_tag=?
               WHERE id=?""",
            (data['meal_date'], data['meal_type'], data['place'], data['amount'],
             data.get('screenshot', ''), data.get('trip_tag', ''), mid),
        )
        conn.commit()
    finally:
        conn.close()


def get_meal(mid):
    conn = get_conn()
    try:
        return conn.execute('SELECT * FROM meals WHERE id=?', (mid,)).fetchone()
    finally:
        conn.close()


def delete_meal(mid):
    conn = get_conn()
    try:
        rec = conn.execute('SELECT screenshot FROM meals WHERE id=?', (mid,)).fetchone()
        if rec and rec['screenshot']:
            remove_screenshot_file(rec['screenshot'])
        conn.execute('DELETE FROM meals WHERE id=?', (mid,))
        conn.commit()
    finally:
        conn.close()


def query_meals(date_from='', date_to='', tag=''):
    sql = 'SELECT * FROM meals WHERE 1=1'
    args = []
    if date_from:
        sql += ' AND meal_date >= ?'
        args.append(date_from)
    if date_to:
        sql += ' AND meal_date <= ?'
        args.append(date_to)
    if tag:
        sql += ' AND trip_tag = ?'
        args.append(tag)
    sql += ' ORDER BY meal_date DESC, id DESC'
    conn = get_conn()
    try:
        return conn.execute(sql, args).fetchall()
    finally:
        conn.close()


def get_meal_place_history():
    conn = get_conn()
    try:
        return [r[0] for r in conn.execute(
            "SELECT place FROM meals WHERE place!='' GROUP BY place ORDER BY MAX(id) DESC LIMIT 50")]
    finally:
        conn.close()


def add_transport(data):
    conn = get_conn()
    try:
        cur = conn.execute(
            """INSERT INTO transports (t_date, mode, depart, arrive, amount, public_flag, screenshot, trip_tag)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (data['t_date'], data['mode'], data['depart'], data['arrive'], data['amount'],
             data['public_flag'], data.get('screenshot', ''), data.get('trip_tag', '')),
        )
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def update_transport(tid, data):
    conn = get_conn()
    try:
        conn.execute(
            """UPDATE transports SET t_date=?, mode=?, depart=?, arrive=?, amount=?, public_flag=?, screenshot=?, trip_tag=?
               WHERE id=?""",
            (data['t_date'], data['mode'], data['depart'], data['arrive'], data['amount'],
             data['public_flag'], data.get('screenshot', ''), data.get('trip_tag', ''), tid),
        )
        conn.commit()
    finally:
        conn.close()


def get_transport(trid):
    conn = get_conn()
    try:
        return conn.execute('SELECT * FROM transports WHERE id=?', (trid,)).fetchone()
    finally:
        conn.close()


def delete_transport(trid):
    conn = get_conn()
    try:
        rec = conn.execute('SELECT screenshot FROM transports WHERE id=?', (trid,)).fetchone()
        if rec and rec['screenshot']:
            remove_screenshot_file(rec['screenshot'])
        conn.execute('DELETE FROM transports WHERE id=?', (trid,))
        conn.commit()
    finally:
        conn.close()


def query_transports(date_from='', date_to='', tag=''):
    sql = 'SELECT * FROM transports WHERE src_trip_id=0'
    args = []
    if date_from:
        sql += ' AND t_date >= ?'
        args.append(date_from)
    if date_to:
        sql += ' AND t_date <= ?'
        args.append(date_to)
    if tag:
        sql += ' AND trip_tag = ?'
        args.append(tag)
    sql += ' ORDER BY t_date DESC, id DESC'
    conn = get_conn()
    try:
        return conn.execute(sql, args).fetchall()
    finally:
        conn.close()


def get_transport_history():
    conn = get_conn()
    try:
        return ([r[0] for r in conn.execute(
            "SELECT depart FROM transports WHERE depart!='' GROUP BY depart ORDER BY MAX(id) DESC LIMIT 50")],
            [r[0] for r in conn.execute(
                "SELECT arrive FROM transports WHERE arrive!='' GROUP BY arrive ORDER BY MAX(id) DESC LIMIT 50")])
    finally:
        conn.close()


def migrate_legacy_trips(conn):
    """把旧 trips 表里的餐饮/交通字段迁到 meals/transports(幂等,两表分别判重)"""
    meals_done = conn.execute('SELECT COUNT(*) FROM meals WHERE src_trip_id>0').fetchone()[0]
    tr_done = conn.execute('SELECT COUNT(*) FROM transports WHERE src_trip_id>0').fetchone()[0]
    if meals_done and tr_done:
        return
    rows = conn.execute('SELECT * FROM trips ORDER BY id').fetchall()
    for r in rows:
        # 交通: 有出行方式即迁移(含费用)
        if r['transport'] and not tr_done:
            conn.execute(
                """INSERT INTO transports
                   (t_date, mode, depart, arrive, amount, public_flag, screenshot, src_trip_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (r['trip_date'], r['transport'], r['depart'], r['arrive'], r['cost'],
                 r['public_flag'], r['public_screenshot'], r['id']),
            )
        # 餐饮: 有标志即迁移
        if r['meal_flag'] and not meals_done:
            conn.execute(
                """INSERT INTO meals
                   (meal_date, meal_type, place, amount, screenshot, src_trip_id)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (r['trip_date'], '', '', 0.0, r['meal_screenshot'], r['id']),
            )


def get_location_history():
    """最近使用过的出发地 / 到达地,用于下拉记忆"""
    conn = get_conn()
    try:
        departs = [r[0] for r in conn.execute(
            "SELECT depart FROM trips WHERE depart!='' GROUP BY depart ORDER BY MAX(id) DESC LIMIT 50")]
        arrives = [r[0] for r in conn.execute(
            "SELECT arrive FROM trips WHERE arrive!='' GROUP BY arrive ORDER BY MAX(id) DESC LIMIT 50")]
        return departs, arrives
    finally:
        conn.close()


def save_screenshot(src_path, trip_date, tag):
    """把选中的图片/PDF复制到 screenshots/日期/ 下,返回相对路径(正斜杠)"""
    day_dir = os.path.join(SCREENSHOTS_DIR, trip_date)
    os.makedirs(day_dir, exist_ok=True)
    ext = os.path.splitext(src_path)[1].lower()
    if ext not in INVOICE_EXTS:
        ext = '.png'
    name = '%s_%s%s' % (tag, datetime.datetime.now().strftime('%H%M%S%f'), ext)
    dest = os.path.join(day_dir, name)
    shutil.copy2(src_path, dest)
    return os.path.relpath(dest, BASE_DIR).replace('\\', '/')


def remove_screenshot_file(rel_path):
    """删除截图文件(忽略错误)"""
    path = rel_to_abs(rel_path)
    if path and os.path.isfile(path):
        try:
            os.remove(path)
        except OSError:
            pass


def get_all_tags():
    """从四张表收集所有已使用的出差标签,去重,按最近使用排序"""
    conn = get_conn()
    try:
        tags = set()
        for tbl in ('trips', 'lodging', 'meals', 'transports'):
            rows = conn.execute(
                "SELECT DISTINCT trip_tag FROM %s WHERE trip_tag IS NOT NULL AND trip_tag != ''" % tbl
            ).fetchall()
            for r in rows:
                tags.add(r[0])
        return sorted(tags, reverse=True)
    finally:
        conn.close()


def calc_trip_total(tag):
    """计算指定标签的出差费用总额, 返回明细 dict"""
    conn = get_conn()
    try:
        # 行程记录金额加总
        r = conn.execute(
            "SELECT COUNT(*), COALESCE(SUM(cost),0) FROM trips WHERE trip_tag=?", (tag,)
        ).fetchone()
        trip_count, trip_total = r[0], r[1]

        # 住宿记录金额加总
        r = conn.execute(
            "SELECT COUNT(*), COALESCE(SUM(amount),0) FROM lodging WHERE trip_tag=?", (tag,)
        ).fetchone()
        lod_count, lod_total = r[0], r[1]

        # 解析城市
        city = parse_city_from_tag(tag)
        tier, meal_rate, transit_rate = classify_city(city)
        tier_names = {1: '一类', 2: '二类', 3: '三类'}

        # 餐饮: 有截图的记录天数 × 餐费标准
        meal_days = conn.execute(
            "SELECT COUNT(DISTINCT meal_date) FROM meals WHERE trip_tag=? AND screenshot IS NOT NULL AND screenshot != ''",
            (tag,)
        ).fetchone()[0]
        meal_cost = meal_days * meal_rate

        # 市内交通: 有截图的记录天数 × 交通标准
        transit_days = conn.execute(
            "SELECT COUNT(DISTINCT t_date) FROM transports WHERE trip_tag=? AND screenshot IS NOT NULL AND screenshot != ''",
            (tag,)
        ).fetchone()[0]
        transit_cost = transit_days * transit_rate

        total = round(trip_total + lod_total + meal_cost + transit_cost, 2)
        return {
            'trip_total': round(trip_total, 2), 'trip_count': trip_count,
            'lod_total': round(lod_total, 2), 'lod_count': lod_count,
            'city': city, 'tier': tier, 'tier_name': tier_names[tier],
            'meal_rate': meal_rate, 'meal_days': meal_days, 'meal_cost': meal_cost,
            'transit_rate': transit_rate, 'transit_days': transit_days, 'transit_cost': transit_cost,
            'total': total,
        }
    finally:
        conn.close()


# ---------------- 导出 Excel ----------------

def export_to_excel(rows):
    """导出全部/筛选记录为 xlsx,返回 (文件路径, 条数)"""
    if not rows:
        return None, 0
    # 批量获取附件计数
    trip_ids = [r['id'] for r in rows]
    file_counts = get_trip_file_counts(trip_ids)
    wb = Workbook()
    ws = wb.active
    ws.title = '行程记录'
    headers = ['序号', '日期', '星期', '出发地', '到达地', '交通方式', '金额(元)', '发票', '行程单', '出差标签']
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='D9E1F2')
    header_font = Font(bold=True)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, r in enumerate(rows, 1):
        fc = file_counts.get(r['id'], {'invoice': 0, 'itinerary': 0})
        inv_text = '有(%d)' % fc['invoice'] if fc['invoice'] else ''
        it_text = '有(%d)' % fc['itinerary'] if fc['itinerary'] else ''
        ws.append([i, r['trip_date'], r['weekday'], r['depart'], r['arrive'],
                   r['transport'], r['cost'], inv_text, it_text, r['trip_tag'] or ''])

    total_cost = round(sum(r['cost'] for r in rows), 2)
    total_row = ws.max_row + 1
    ws.append(['', '', '', '', '', '总条数', total_cost, len(rows), '', ''])

    # 表头样式
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 数据样式
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            if cell.column == 7:  # 金额列
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')

    # 合计行样式
    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='FCE4D6')

    widths = [6, 12, 9, 16, 16, 14, 12, 10, 10]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = w
    ws.freeze_panes = 'A2'

    os.makedirs(EXPORT_DIR, exist_ok=True)
    path = os.path.join(EXPORT_DIR, '差旅费用记录_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d_%H%M%S'))
    wb.save(path)
    return path, len(rows)


def export_lodging_to_excel(rows):
    """导出全部住宿记录为 xlsx,返回 (文件路径, 条数)"""
    if not rows:
        return None, 0
    # 批量获取附件计数
    lod_ids = [r['id'] for r in rows]
    file_counts = get_lodging_file_counts(lod_ids)
    wb = Workbook()
    ws = wb.active
    ws.title = '住宿记录'
    headers = ['序号', '入住日期', '退房日期', '酒店名称', '金额(元)', '发票', '水单', '出差标签']
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='D9E1F2')
    header_font = Font(bold=True)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, r in enumerate(rows, 1):
        fc = file_counts.get(r['id'], {'invoice': 0, 'receipt': 0})
        inv_text = '有(%d)' % fc['invoice'] if fc['invoice'] else ''
        rec_text = '有(%d)' % fc['receipt'] if fc['receipt'] else ''
        ws.append([i, r['checkin_date'], r['checkout_date'], r['hotel'], r['amount'],
                   inv_text, rec_text, r['trip_tag'] or ''])

    total = sum(r['amount'] for r in rows)
    total_row = ws.max_row + 1
    ws.append(['', '', '', '合计', round(total, 2), '', ''])

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            if cell.column == 5:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')

    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='FCE4D6')

    widths = [6, 12, 12, 30, 12, 10, 10]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = w
    ws.freeze_panes = 'A2'

    os.makedirs(EXPORT_DIR, exist_ok=True)
    path = os.path.join(EXPORT_DIR, '住宿记录_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d_%H%M%S'))
    wb.save(path)
    return path, len(rows)


def _style_workbook(ws, rows, headers, total_text):
    """通用: 写入表头+数据+合计, 返回 total_row"""
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    header_font = Font(bold=True)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    total_row = ws.max_row + 1
    ws.append(total_text)
    for cell in ws[total_row]:
        cell.font = Font(bold=True)
    return total_row


def export_meals_to_excel(rows):
    if not rows:
        return None, 0
    wb = Workbook()
    ws = wb.active
    ws.title = '餐饮记录'
    headers = ['序号', '日期', '餐次', '用餐截图', '出差标签']
    for i, r in enumerate(rows, 1):
        ws.append([i, r['meal_date'], r['meal_type'], r['screenshot'], r['trip_tag'] or ''])
    _style_workbook(ws, rows, headers, ['', '', '共 %d 条' % len(rows), '', ''])
    ws.freeze_panes = 'A2'
    for idx, w in enumerate([6, 12, 10, 40, 30], 1):
        ws.column_dimensions[chr(64 + idx)].width = w
    path = os.path.join(EXPORT_DIR, '餐饮记录_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d_%H%M%S'))
    wb.save(path)
    return path, len(rows)


def export_transports_to_excel(rows):
    if not rows:
        return None, 0
    wb = Workbook()
    ws = wb.active
    ws.title = '交通记录'
    headers = ['序号', '日期', '出发地', '到达地', '金额(元)', '交通截图', '出差标签']
    for i, r in enumerate(rows, 1):
        ws.append([i, r['t_date'], r['depart'], r['arrive'], r['amount'], r['screenshot'], r['trip_tag'] or ''])
    total = round(sum(r['amount'] for r in rows), 2)
    _style_workbook(ws, rows, headers, ['', '', '', '合计', total, '', ''])
    ws.freeze_panes = 'A2'
    for idx, w in enumerate([6, 12, 16, 16, 12, 40, 30], 1):
        ws.column_dimensions[chr(64 + idx)].width = w
    path = os.path.join(EXPORT_DIR, '交通记录_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d_%H%M%S'))
    wb.save(path)
    return path, len(rows)


# ---------------- 行程单 PDF 解析 ----------------

def parse_itinerary_pdf(pdf_path):
    """解析高德打车行程单 PDF,返回 trip 列表:
    [{'date': str, 'depart': str, 'arrive': str, 'cost': float}, ...]
    使用 PyMuPDF dict 模式按列 x 坐标精确区分起点/终点。
    """
    try:
        doc = fitz.open(pdf_path)
        page_data = doc[0].get_text('dict')
        doc.close()
    except Exception:
        return []

    _DT = re.compile(r'\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}')
    _AMT = re.compile(r'(\d+\.?\d*)元')

    # 收集每个文本块的 span 列表: [(x0, text), ...]
    def _block_spans(block):
        result = []
        for line in block.get('lines', []):
            for span in line.get('spans', []):
                t = span['text'].strip()
                if t:
                    result.append((span['bbox'][0], t))
        return result

    blocks = [b for b in page_data.get('blocks', []) if b.get('type') == 0]

    # 1. 找到表头块, 提取列 x 坐标用于精确定位
    city_x = None
    depart_x = None
    dest_x = None
    hdr_idx = -1
    for idx, block in enumerate(blocks):
        spans = _block_spans(block)
        if any(t == '序号' for _, t in spans):
            for x, t in spans:
                if t == '城市':
                    city_x = x
                elif t == '起点':
                    depart_x = x
                elif t == '终点':
                    dest_x = x
            hdr_idx = idx
            break
    if depart_x is None or dest_x is None or hdr_idx < 0:
        return []
    # 用表头列坐标的中点作为数据分类边界
    left_x = (city_x + depart_x) / 2 if city_x else depart_x - 30
    mid_x = (depart_x + dest_x) / 2

    # 2. 解析表头之后的数据块
    trips = []
    for block in blocks[hdr_idx + 1:]:
        spans = _block_spans(block)
        if not spans or not spans[0][1].isdigit():
            continue

        # 提取上车时间
        date_str = None
        for _, t in spans:
            m = _DT.search(t)
            if m:
                date_str = m.group()[:10]
                break

        # 按 x 坐标拆分起点/终点, 提取金额
        depart_parts = []
        arrive_parts = []
        cost = None
        for x, t in spans:
            if _AMT.match(t):
                cost = float(_AMT.match(t).group(1))
            elif x >= mid_x:
                arrive_parts.append(t)
            elif x >= left_x:  # 起点列
                depart_parts.append(t)

        if date_str and depart_parts and cost is not None:
            trips.append({
                'date': date_str,
                'depart': ''.join(depart_parts),
                'arrive': ''.join(arrive_parts),
                'cost': cost,
            })

    return trips


# ---------------- 新增/编辑对话框 ----------------

class EditDialog(tk.Toplevel):
    """新增或编辑一条行程记录（含金额 + 发票上传）"""

    def __init__(self, master, record=None):
        super().__init__(master)
        self.record = record or {}
        is_edit = bool(self.record)
        self.title('编辑行程' if is_edit else '新增行程')
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        today = datetime.date.today().isoformat()
        self.date_var = tk.StringVar(value=self.record.get('trip_date') or today)
        self.weekday_var = tk.StringVar()
        self.depart_var = tk.StringVar(value=self.record.get('depart') or '')
        self.arrive_var = tk.StringVar(value=self.record.get('arrive') or '')
        self.transport_var = tk.StringVar(value=self.record.get('transport') or '')
        cost = self.record.get('cost')
        self.cost_var = tk.StringVar(value=('%g' % cost) if cost else '')
        self.tag_var = tk.StringVar(value=self.record.get('trip_tag') or '')

        # 附件: {id: {type, path, name}} 已入库文件
        self.att_files = {}
        # 待保存新文件: [(file_type, src_abs_path)]
        self.pending_files = []
        # 已删除的文件 id(编辑时)
        self.deleted_fids = []
        if is_edit:
            for f in get_trip_files(self.record['id']):
                self.att_files[f['id']] = {
                    'type': f['file_type'], 'path': f['file_path'], 'name': f['file_name'],
                }

        self._build()
        self._update_weekday()
        self._refresh_file_lists()

        # 居中显示
        self.update_idletasks()
        x = master.winfo_rootx() + (master.winfo_width() - self.winfo_width()) // 2
        y = master.winfo_rooty() + (master.winfo_height() - self.winfo_height()) // 2
        self.geometry('+%d+%d' % (max(x, 0), max(y, 0)))

    # ---------- 界面构建 ----------
    def _build(self):
        pad = {'padx': 10, 'pady': 4}
        body = ttk.Frame(self, padding=12)
        body.pack(fill='both', expand=True)

        # 日期
        row = ttk.Frame(body)
        row.pack(fill='x', **pad)
        ttk.Label(row, text='日期 *').pack(side='left')
        ttk.Entry(row, textvariable=self.date_var, width=14).pack(side='left', padx=(6, 4))
        ttk.Button(row, text='今天', width=6, command=self._set_today).pack(side='left')
        ttk.Label(row, textvariable=self.weekday_var, font=FONT_BOLD).pack(side='left', padx=(10, 0))

        # 出发地
        row = ttk.Frame(body)
        row.pack(fill='x', **pad)
        ttk.Label(row, text='出发地 *').pack(side='left')
        departs, arrives = get_location_history()
        cbo = ttk.Combobox(row, textvariable=self.depart_var, values=departs, width=30)
        cbo.pack(side='left', padx=(6, 0))

        # 到达地
        row = ttk.Frame(body)
        row.pack(fill='x', **pad)
        ttk.Label(row, text='到达地 *').pack(side='left')
        ttk.Combobox(row, textvariable=self.arrive_var, values=arrives, width=30).pack(side='left', padx=(6, 0))

        # 交通方式
        row = ttk.Frame(body)
        row.pack(fill='x', **pad)
        ttk.Label(row, text='交通方式 *').pack(side='left')
        ttk.Combobox(row, textvariable=self.transport_var, values=TRANSPORT_OPTIONS,
                     state='readonly', width=28).pack(side='left', padx=(6, 0))

        # 金额
        row = ttk.Frame(body)
        row.pack(fill='x', **pad)
        ttk.Label(row, text='金额(元)').pack(side='left')
        ttk.Entry(row, textvariable=self.cost_var, width=12).pack(side='left', padx=(6, 0))

        # --- 发票 ---
        inv_box = ttk.LabelFrame(body, text='发票', padding=6)
        inv_box.pack(fill='x', **pad)
        inv_top = ttk.Frame(inv_box)
        inv_top.pack(fill='x')
        ttk.Button(inv_top, text='上传发票', width=10,
                   command=lambda: self._pick_files('invoice')).pack(side='left')
        self.inv_count_lbl = ttk.Label(inv_top, text='', foreground='#666666')
        self.inv_count_lbl.pack(side='left', padx=(10, 0))
        self.inv_list_frame = ttk.Frame(inv_box)
        self.inv_list_frame.pack(fill='x', pady=(4, 0))

        # --- 电子行程单 ---
        it_box = ttk.LabelFrame(body, text='电子行程单', padding=6)
        it_box.pack(fill='x', **pad)
        it_top = ttk.Frame(it_box)
        it_top.pack(fill='x')
        ttk.Button(it_top, text='上传行程单', width=10,
                   command=lambda: self._pick_files('itinerary')).pack(side='left')
        ttk.Button(it_top, text='识别行程单', width=10,
                   command=self._ocr_itinerary).pack(side='left', padx=(6, 0))
        self.it_count_lbl = ttk.Label(it_top, text='', foreground='#666666')
        self.it_count_lbl.pack(side='left', padx=(10, 0))
        self.it_list_frame = ttk.Frame(it_box)
        self.it_list_frame.pack(fill='x', pady=(4, 0))

        # 出差标签
        row = ttk.Frame(body)
        row.pack(fill='x', **pad)
        ttk.Label(row, text='出差标签').pack(side='left')
        ttk.Combobox(row, textvariable=self.tag_var, values=get_all_tags(), width=30).pack(side='left', padx=(6, 0))

        # 按钮行
        row = ttk.Frame(body)
        row.pack(fill='x', pady=(10, 0))
        ttk.Button(row, text='保存', width=10, command=self._save).pack(side='right', padx=(8, 0))
        ttk.Button(row, text='取消', width=10, command=self.destroy).pack(side='right')

    # ---------- 行为 ----------
    def _set_today(self):
        self.date_var.set(datetime.date.today().isoformat())
        self._update_weekday()

    def _update_weekday(self):
        d = parse_date(self.date_var.get())
        self.weekday_var.set(weekday_of(self.date_var.get()) if d else '日期格式: YYYY-MM-DD')

    def _pick_files(self, file_type):
        """选择多个文件"""
        title = '选择发票' if file_type == 'invoice' else '选择电子行程单'
        paths = filedialog.askopenfilenames(
            title=title, parent=self,
            filetypes=[('发票文件', '*.png *.jpg *.jpeg *.bmp *.gif *.webp *.pdf'),
                       ('图片文件', '*.png *.jpg *.jpeg *.bmp *.gif *.webp'),
                       ('PDF文件', '*.pdf'), ('所有文件', '*.*')])
        if not paths:
            return
        for p in paths:
            self.pending_files.append((file_type, p))
        self._refresh_file_lists()

    def _refresh_file_lists(self):
        """刷新两个附件列表"""
        for ft, frame, count_lbl, tag in [
            ('invoice', self.inv_list_frame, self.inv_count_lbl, '发票'),
            ('itinerary', self.it_list_frame, self.it_count_lbl, '行程单'),
        ]:
            for w in frame.winfo_children():
                w.destroy()
            items = [(fid, v) for fid, v in self.att_files.items() if v['type'] == ft]
            pending = [(i, v) for i, v in enumerate(self.pending_files) if v[0] == ft]
            count = len(items) + len(pending)
            count_lbl.configure(text='%d 个文件' % count if count else '未上传')
            for fid, v in items:
                row = ttk.Frame(frame)
                row.pack(fill='x', pady=1)
                name = v['name'] or os.path.basename(v['path'])
                if len(name) > 30:
                    name = name[:27] + '...'
                ttk.Label(row, text=name, foreground='#333333').pack(side='left')
                ttk.Button(row, text='移除', width=4,
                           command=lambda fid=fid: self._remove_existing(fid)).pack(side='left', padx=(8, 0))
            for pi, (ft2, src) in enumerate(self.pending_files):
                if ft2 != ft:
                    continue
                row = ttk.Frame(frame)
                row.pack(fill='x', pady=1)
                name = os.path.basename(src)
                if len(name) > 30:
                    name = name[:27] + '...'
                ttk.Label(row, text='[新] ' + name, foreground='#0563C1').pack(side='left')
                ttk.Button(row, text='移除', width=4,
                           command=lambda pi=pi: self._remove_pending(pi)).pack(side='left', padx=(8, 0))

    def _remove_existing(self, fid):
        """标记已入库文件为删除"""
        self.deleted_fids.append(fid)
        del self.att_files[fid]
        self._refresh_file_lists()

    def _remove_pending(self, idx):
        """移除待保存文件"""
        if 0 <= idx < len(self.pending_files):
            self.pending_files.pop(idx)
        self._refresh_file_lists()

    def _ocr_itinerary(self):
        """从已上传的行程单 PDF 中解析打车信息,自动填充表单"""
        pdfs = [(i, src) for i, (ft, src) in enumerate(self.pending_files)
                if ft == 'itinerary' and src.lower().endswith('.pdf')]
        if not pdfs:
            messagebox.showinfo('提示', '请先上传行程单 PDF 文件', parent=self)
            return

        all_trips = []
        for _, src in pdfs:
            all_trips.extend(parse_itinerary_pdf(src))

        if not all_trips:
            messagebox.showwarning('提示', '未能从行程单中识别到有效行程信息', parent=self)
            return

        t = all_trips[0]
        self.date_var.set(t['date'])
        self._update_weekday()
        self.depart_var.set(t['depart'])
        self.arrive_var.set(t['arrive'])
        self.transport_var.set('打车')
        self.cost_var.set('%g' % t['cost'])

        n = len(all_trips)
        msg = '已识别并填充行程信息' + (' (共 %d 条,已填入第 1 条)' % n if n > 1 else '')
        messagebox.showinfo('识别成功', msg, parent=self)

    def _save(self):
        date_str = self.date_var.get().strip()
        d = parse_date(date_str)
        if d is None:
            messagebox.showwarning('提示', '日期格式不正确,应为 YYYY-MM-DD,例如 2026-08-17', parent=self)
            return
        depart = self.depart_var.get().strip()
        arrive = self.arrive_var.get().strip()
        transport = self.transport_var.get().strip()
        if not depart or not arrive or not transport:
            messagebox.showwarning('提示', '请填写出发地、到达地、交通方式', parent=self)
            return
        # 金额
        cost = 0.0
        cost_text = self.cost_var.get().strip()
        if cost_text:
            try:
                cost = round(float(cost_text), 2)
            except ValueError:
                messagebox.showwarning('提示', '金额请输入数字', parent=self)
                return

        date_iso = d.isoformat()
        data = {
            'trip_date': date_iso,
            'weekday': weekday_of(date_iso),
            'depart': depart,
            'arrive': arrive,
            'transport': transport,
            'cost': cost,
            'meal_flag': 0,
            'meal_screenshot': '',
            'public_flag': 0,
            'public_screenshot': '',
            'trip_tag': self.tag_var.get().strip(),
        }

        if self.record:
            # 编辑: 直接操作数据库
            tid = self.record['id']
            update_trip(tid, data)
            # 删除已标记的文件
            for fid in self.deleted_fids:
                delete_trip_file(fid, delete_file=True)
            # 保存新文件
            for ft, src in self.pending_files:
                rel = save_screenshot(src, date_iso, ft)
                add_trip_file(tid, ft, rel, os.path.basename(src))
        else:
            # 新增: 先创建行程再保存附件
            tid = add_trip(data)
            for ft, src in self.pending_files:
                rel = save_screenshot(src, date_iso, ft)
                add_trip_file(tid, ft, rel, os.path.basename(src))
        self.destroy()


class LodgingDialog(tk.Toplevel):
    """新增或编辑一条住宿记录"""

    def __init__(self, master, record=None):
        super().__init__(master)
        self.record = record or {}
        is_edit = bool(self.record)
        self.title('编辑住宿' if is_edit else '新增住宿')
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        today = datetime.date.today().isoformat()
        self.checkin_var = tk.StringVar(value=self.record.get('checkin_date') or today)
        self.checkout_var = tk.StringVar(value=self.record.get('checkout_date') or today)
        self.hotel_var = tk.StringVar(value=self.record.get('hotel') or '')
        amt = self.record.get('amount')
        self.amount_var = tk.StringVar(value=('%g' % amt) if amt else '')
        self.tag_var = tk.StringVar(value=self.record.get('trip_tag') or '')

        # 附件: {id: {type, path, name}}
        self.att_files = {}
        self.pending_files = []
        self.deleted_fids = []
        if is_edit:
            for f in get_lodging_files(self.record['id']):
                self.att_files[f['id']] = {
                    'type': f['file_type'], 'path': f['file_path'], 'name': f['file_name'],
                }

        self._build()
        self._refresh_file_lists()
        self.update_idletasks()
        x = master.winfo_rootx() + (master.winfo_width() - self.winfo_width()) // 2
        y = master.winfo_rooty() + (master.winfo_height() - self.winfo_height()) // 2
        self.geometry('+%d+%d' % (max(x, 0), max(y, 0)))

    def _build(self):
        pad = {'padx': 10, 'pady': 4}
        body = ttk.Frame(self, padding=14)
        body.pack(fill='both', expand=True)

        rows = [
            ('入住日期 *', self.checkin_var, 'YYYY-MM-DD'),
            ('退房日期 *', self.checkout_var, 'YYYY-MM-DD'),
        ]
        hotel_names = get_lodging_history()
        for text, var, ph in rows:
            row = ttk.Frame(body)
            row.pack(fill='x', **pad)
            ttk.Label(row, text=text).pack(side='left')
            ttk.Entry(row, textvariable=var, width=22).pack(side='left', padx=(6, 0))
            ttk.Label(row, text='(%s)' % ph, foreground='#999999').pack(side='left', padx=(6, 0))

        # 酒店名称(带记忆下拉)
        row = ttk.Frame(body)
        row.pack(fill='x', **pad)
        ttk.Label(row, text='酒店名称 *').pack(side='left')
        ttk.Combobox(row, textvariable=self.hotel_var, values=hotel_names, width=30).pack(side='left', padx=(6, 0))

        # 金额
        row = ttk.Frame(body)
        row.pack(fill='x', **pad)
        ttk.Label(row, text='金额(元) *').pack(side='left')
        ttk.Entry(row, textvariable=self.amount_var, width=12).pack(side='left', padx=(6, 0))

        # --- 发票 ---
        inv_box = ttk.LabelFrame(body, text='发票', padding=6)
        inv_box.pack(fill='x', **pad)
        inv_top = ttk.Frame(inv_box)
        inv_top.pack(fill='x')
        ttk.Button(inv_top, text='上传发票', width=10,
                   command=lambda: self._pick_files('invoice')).pack(side='left')
        self.inv_count_lbl = ttk.Label(inv_top, text='', foreground='#666666')
        self.inv_count_lbl.pack(side='left', padx=(10, 0))
        self.inv_list_frame = ttk.Frame(inv_box)
        self.inv_list_frame.pack(fill='x', pady=(4, 0))

        # --- 水单 ---
        rec_box = ttk.LabelFrame(body, text='水单', padding=6)
        rec_box.pack(fill='x', **pad)
        rec_top = ttk.Frame(rec_box)
        rec_top.pack(fill='x')
        ttk.Button(rec_top, text='上传水单', width=10,
                   command=lambda: self._pick_files('receipt')).pack(side='left')
        self.rec_count_lbl = ttk.Label(rec_top, text='', foreground='#666666')
        self.rec_count_lbl.pack(side='left', padx=(10, 0))
        self.rec_list_frame = ttk.Frame(rec_box)
        self.rec_list_frame.pack(fill='x', pady=(4, 0))

        # 出差标签
        row = ttk.Frame(body)
        row.pack(fill='x', **pad)
        ttk.Label(row, text='出差标签').pack(side='left')
        ttk.Combobox(row, textvariable=self.tag_var, values=get_all_tags(), width=30).pack(side='left', padx=(6, 0))

        # 按钮行
        row = ttk.Frame(body)
        row.pack(fill='x', pady=(12, 0))
        ttk.Button(row, text='保存', width=10, command=self._save).pack(side='right', padx=(8, 0))
        ttk.Button(row, text='取消', width=10, command=self.destroy).pack(side='right')

    # ---------- 附件操作 ----------
    def _pick_files(self, file_type):
        title = '选择发票' if file_type == 'invoice' else '选择水单'
        paths = filedialog.askopenfilenames(
            title=title, parent=self,
            filetypes=[('发票文件', '*.png *.jpg *.jpeg *.bmp *.gif *.webp *.pdf'),
                       ('图片文件', '*.png *.jpg *.jpeg *.bmp *.gif *.webp'),
                       ('PDF文件', '*.pdf'), ('所有文件', '*.*')])
        if not paths:
            return
        for p in paths:
            self.pending_files.append((file_type, p))
        self._refresh_file_lists()

    def _refresh_file_lists(self):
        for ft, frame, count_lbl, tag in [
            ('invoice', self.inv_list_frame, self.inv_count_lbl, '发票'),
            ('receipt', self.rec_list_frame, self.rec_count_lbl, '水单'),
        ]:
            for w in frame.winfo_children():
                w.destroy()
            items = [(fid, v) for fid, v in self.att_files.items() if v['type'] == ft]
            pending = [(pi, v) for pi, v in enumerate(self.pending_files) if v[0] == ft]
            count = len(items) + len(pending)
            count_lbl.configure(text='%d 个文件' % count if count else '未上传')
            for fid, v in items:
                row = ttk.Frame(frame)
                row.pack(fill='x', pady=1)
                name = v['name'] or os.path.basename(v['path'])
                if len(name) > 30:
                    name = name[:27] + '...'
                ttk.Label(row, text=name, foreground='#333333').pack(side='left')
                ttk.Button(row, text='移除', width=4,
                           command=lambda fid=fid: self._remove_existing(fid)).pack(side='left', padx=(8, 0))
            for pi, (ft2, src) in enumerate(self.pending_files):
                if ft2 != ft:
                    continue
                row = ttk.Frame(frame)
                row.pack(fill='x', pady=1)
                name = os.path.basename(src)
                if len(name) > 30:
                    name = name[:27] + '...'
                ttk.Label(row, text='[新] ' + name, foreground='#0563C1').pack(side='left')
                ttk.Button(row, text='移除', width=4,
                           command=lambda pi=pi: self._remove_pending(pi)).pack(side='left', padx=(8, 0))

    def _remove_existing(self, fid):
        self.deleted_fids.append(fid)
        del self.att_files[fid]
        self._refresh_file_lists()

    def _remove_pending(self, idx):
        if 0 <= idx < len(self.pending_files):
            self.pending_files.pop(idx)
        self._refresh_file_lists()

    def _save(self):
        cin = self.checkin_var.get().strip()
        cout = self.checkout_var.get().strip()
        hotel = self.hotel_var.get().strip()
        if parse_date(cin) is None or parse_date(cout) is None:
            messagebox.showwarning('提示', '日期格式不正确,应为 YYYY-MM-DD', parent=self)
            return
        if not hotel:
            messagebox.showwarning('提示', '请填写酒店名称', parent=self)
            return
        if parse_date(cout) < parse_date(cin):
            messagebox.showwarning('提示', '退房日期不能早于入住日期', parent=self)
            return
        text = self.amount_var.get().strip()
        if not text:
            messagebox.showwarning('提示', '请填写金额', parent=self)
            return
        try:
            amount = round(float(text), 2)
        except ValueError:
            messagebox.showwarning('提示', '金额请输入数字', parent=self)
            return
        data = {
            'checkin_date': parse_date(cin).isoformat(),
            'checkout_date': parse_date(cout).isoformat(),
            'hotel': hotel,
            'amount': amount,
            'trip_tag': self.tag_var.get().strip(),
        }
        if self.record:
            lid = self.record['id']
            update_lodging(lid, data)
            for fid in self.deleted_fids:
                delete_lodging_file(fid, delete_file=True)
            for ft, src in self.pending_files:
                rel = save_screenshot(src, data['checkin_date'], 'lod_' + ft)
                add_lodging_file(lid, ft, rel, os.path.basename(src))
        else:
            lid = add_lodging(data)
            for ft, src in self.pending_files:
                rel = save_screenshot(src, data['checkin_date'], 'lod_' + ft)
                add_lodging_file(lid, ft, rel, os.path.basename(src))
        self.destroy()


class MealDialog(tk.Toplevel):
    """新增或编辑一条餐饮记录（仅记录餐次 + 截图凭证）"""

    def __init__(self, master, record=None):
        super().__init__(master)
        self.record = record or {}
        self.title('编辑餐饮' if self.record else '新增餐饮')
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        today = datetime.date.today().isoformat()
        self.date_var = tk.StringVar(value=self.record.get('meal_date') or today)
        self.type_var = tk.StringVar(value=self.record.get('meal_type') or '午餐')
        self.tag_var = tk.StringVar(value=self.record.get('trip_tag') or '')

        # 截图: 编辑时已入库的相对路径 / 新选择的源路径
        self.old_shot = self.record.get('screenshot') or ''
        self.new_src = ''
        self.thumb = None

        self._build()
        if self.old_shot:
            self._preview_shot(rel_to_abs(self.old_shot))
        self.update_idletasks()
        x = master.winfo_rootx() + (master.winfo_width() - self.winfo_width()) // 2
        y = master.winfo_rooty() + (master.winfo_height() - self.winfo_height()) // 2
        self.geometry('+%d+%d' % (max(x, 0), max(y, 0)))

    def _build(self):
        pad = {'padx': 10, 'pady': 4}
        body = ttk.Frame(self, padding=14)
        body.pack(fill='both', expand=True)

        row = ttk.Frame(body); row.pack(fill='x', **pad)
        ttk.Label(row, text='日期 *').pack(side='left')
        ttk.Entry(row, textvariable=self.date_var, width=14).pack(side='left', padx=(6, 0))

        row = ttk.Frame(body); row.pack(fill='x', **pad)
        ttk.Label(row, text='餐次 *').pack(side='left')
        ttk.Combobox(row, textvariable=self.type_var, values=MEAL_TYPES,
                     state='readonly', width=12).pack(side='left', padx=(6, 0))

        # 用餐截图
        shot_box = ttk.LabelFrame(body, text='用餐截图', padding=8)
        shot_box.pack(fill='x', **pad)
        self.shot_btn = ttk.Button(shot_box, text='选择截图', width=10, command=self._pick_shot)
        self.shot_btn.pack(side='left')
        self.shot_label = ttk.Label(shot_box, text='未选择截图', foreground='#999999')
        self.shot_label.pack(side='left', padx=(10, 0))

        # 出差标签
        row = ttk.Frame(body); row.pack(fill='x', **pad)
        ttk.Label(row, text='出差标签').pack(side='left')
        ttk.Combobox(row, textvariable=self.tag_var, values=get_all_tags(), width=30).pack(side='left', padx=(6, 0))

        row = ttk.Frame(body); row.pack(fill='x', pady=(12, 0))
        ttk.Button(row, text='保存', width=10, command=self._save).pack(side='right', padx=(8, 0))
        ttk.Button(row, text='取消', width=10, command=self.destroy).pack(side='right')

    def _pick_shot(self):
        path = filedialog.askopenfilename(
            title='选择用餐截图', parent=self,
            filetypes=[('图片文件', '*.png *.jpg *.jpeg *.bmp *.gif *.webp'), ('所有文件', '*.*')])
        if not path:
            return
        self.new_src = path
        self._preview_shot(path)

    def _preview_shot(self, path):
        try:
            img = Image.open(path)
            img.thumbnail((150, 100))
            photo = ImageTk.PhotoImage(img)
            self.shot_label.configure(image=photo, text='')
            self.thumb = photo
        except Exception:
            self.shot_label.configure(image='', text='无法预览该图片')

    def _save(self):
        d = parse_date(self.date_var.get().strip())
        if d is None:
            messagebox.showwarning('提示', '日期格式不正确,应为 YYYY-MM-DD', parent=self)
            return
        # 截图处理
        if self.new_src:
            shot = save_screenshot(self.new_src, d.isoformat(), 'meal')
            if self.old_shot:
                remove_screenshot_file(self.old_shot)
        elif self.old_shot:
            shot = self.old_shot
        else:
            shot = ''
        data = {
            'meal_date': d.isoformat(),
            'meal_type': self.type_var.get().strip() or '其他',
            'place': '',
            'amount': 0.0,
            'screenshot': shot,
            'trip_tag': self.tag_var.get().strip(),
        }
        if self.record:
            update_meal(self.record['id'], data)
        else:
            add_meal(data)
        self.destroy()


class TransportDialog(tk.Toplevel):
    """新增或编辑一条交通记录（日期/出发地/到达地/金额/截图）"""

    def __init__(self, master, record=None):
        super().__init__(master)
        self.record = record or {}
        self.title('编辑交通' if self.record else '新增交通')
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        today = datetime.date.today().isoformat()
        self.date_var = tk.StringVar(value=self.record.get('t_date') or today)
        departs, arrives = get_transport_history()
        self.depart_var = tk.StringVar(value=self.record.get('depart') or '')
        self.arrive_var = tk.StringVar(value=self.record.get('arrive') or '')
        amt = self.record.get('amount')
        self.amount_var = tk.StringVar(value=('%g' % amt) if amt else '')
        self._departs, self._arrives = departs, arrives
        self.tag_var = tk.StringVar(value=self.record.get('trip_tag') or '')
        # 截图: 编辑时已入库的相对路径 / 新选择的源路径
        self.old_shot = self.record.get('screenshot') or ''
        self.new_src = ''
        self.thumb = None

        self._build()
        if self.old_shot:
            self._preview_shot(rel_to_abs(self.old_shot))
        self.update_idletasks()
        x = master.winfo_rootx() + (master.winfo_width() - self.winfo_width()) // 2
        y = master.winfo_rooty() + (master.winfo_height() - self.winfo_height()) // 2
        self.geometry('+%d+%d' % (max(x, 0), max(y, 0)))

    def _build(self):
        pad = {'padx': 10, 'pady': 4}
        body = ttk.Frame(self, padding=14)
        body.pack(fill='both', expand=True)

        row = ttk.Frame(body); row.pack(fill='x', **pad)
        ttk.Label(row, text='日期 *').pack(side='left')
        ttk.Entry(row, textvariable=self.date_var, width=14).pack(side='left', padx=(6, 0))

        row = ttk.Frame(body); row.pack(fill='x', **pad)
        ttk.Label(row, text='出发地').pack(side='left')
        ttk.Combobox(row, textvariable=self.depart_var, values=self._departs, width=28).pack(side='left', padx=(6, 0))

        row = ttk.Frame(body); row.pack(fill='x', **pad)
        ttk.Label(row, text='到达地').pack(side='left')
        ttk.Combobox(row, textvariable=self.arrive_var, values=self._arrives, width=28).pack(side='left', padx=(6, 0))

        row = ttk.Frame(body); row.pack(fill='x', **pad)
        ttk.Label(row, text='金额(元) *').pack(side='left')
        ttk.Entry(row, textvariable=self.amount_var, width=12).pack(side='left', padx=(6, 0))

        # 交通截图
        shot_box = ttk.LabelFrame(body, text='交通截图', padding=8)
        shot_box.pack(fill='x', **pad)
        self.shot_btn = ttk.Button(shot_box, text='选择截图', width=10, command=self._pick_shot)
        self.shot_btn.pack(side='left')
        self.shot_label = ttk.Label(shot_box, text='未选择截图', foreground='#999999')
        self.shot_label.pack(side='left', padx=(10, 0))

        # 出差标签
        row = ttk.Frame(body); row.pack(fill='x', **pad)
        ttk.Label(row, text='出差标签').pack(side='left')
        ttk.Combobox(row, textvariable=self.tag_var, values=get_all_tags(), width=30).pack(side='left', padx=(6, 0))

        row = ttk.Frame(body); row.pack(fill='x', pady=(12, 0))
        ttk.Button(row, text='保存', width=10, command=self._save).pack(side='right', padx=(8, 0))
        ttk.Button(row, text='取消', width=10, command=self.destroy).pack(side='right')

    def _pick_shot(self):
        path = filedialog.askopenfilename(
            title='选择交通截图', parent=self,
            filetypes=[('图片文件', '*.png *.jpg *.jpeg *.bmp *.gif *.webp'), ('所有文件', '*.*')])
        if not path:
            return
        self.new_src = path
        self._preview_shot(path)

    def _preview_shot(self, path):
        try:
            img = Image.open(path)
            img.thumbnail((150, 100))
            photo = ImageTk.PhotoImage(img)
            self.shot_label.configure(image=photo, text='')
            self.thumb = photo
        except Exception:
            self.shot_label.configure(image='', text='无法预览该图片')

    def _save(self):
        d = parse_date(self.date_var.get().strip())
        if d is None:
            messagebox.showwarning('提示', '日期格式不正确,应为 YYYY-MM-DD', parent=self)
            return
        text = self.amount_var.get().strip()
        if not text:
            messagebox.showwarning('提示', '请填写金额', parent=self)
            return
        try:
            amount = round(float(text), 2)
        except ValueError:
            messagebox.showwarning('提示', '金额请输入数字', parent=self)
            return
        # 截图处理
        if self.new_src:
            shot = save_screenshot(self.new_src, d.isoformat(), 'public')
            if self.old_shot:
                remove_screenshot_file(self.old_shot)
        elif self.old_shot:
            shot = self.old_shot
        else:
            shot = ''
        data = {
            't_date': d.isoformat(),
            'mode': '',
            'depart': self.depart_var.get().strip(),
            'arrive': self.arrive_var.get().strip(),
            'amount': amount,
            'public_flag': 0,
            'screenshot': shot,
            'trip_tag': self.tag_var.get().strip(),
        }
        if self.record:
            update_transport(self.record['id'], data)
        else:
            add_transport(data)
        self.destroy()


# ---------------- 主窗口 ----------------

class MainApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title('差旅费用详情记录')
        self.geometry('1120x640')
        self.minsize(900, 500)

        style = ttk.Style(self)
        style.configure('.', font=FONT)
        style.configure('Treeview', rowheight=26, font=FONT)
        style.configure('Treeview.Heading', font=FONT_BOLD)

        self.date_from_var = tk.StringVar()
        self.date_to_var = tk.StringVar()
        self.keyword_var = tk.StringVar()

        # 餐饮/交通日期筛选变量
        self.meal_date_from_var = tk.StringVar()
        self.meal_date_to_var = tk.StringVar()
        self.tr_date_from_var = tk.StringVar()
        self.tr_date_to_var = tk.StringVar()

        # 标签筛选变量
        self.trip_tag_var = tk.StringVar()
        self.lod_tag_var = tk.StringVar()
        self.meal_tag_var = tk.StringVar()
        self.tr_tag_var = tk.StringVar()

        self._build_notebook()
        self._build_calc_panel()
        self._build_statusbar()

        # 默认显示本月记录
        today = datetime.date.today()
        first_of_month = today.replace(day=1).isoformat()
        today_str = today.isoformat()
        self.date_from_var.set(first_of_month)
        self.date_to_var.set(today_str)
        self.meal_date_from_var.set(first_of_month)
        self.meal_date_to_var.set(today_str)
        self.tr_date_from_var.set(first_of_month)
        self.tr_date_to_var.set(today_str)
        self.refresh()
        self.refresh_lodging()
        self.refresh_meals()
        self.refresh_transports()

    # ---------- 界面构建 ----------
    def _build_notebook(self):
        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True)

        self.trip_tab = ttk.Frame(nb, padding=0)
        self.lod_tab = ttk.Frame(nb, padding=0)
        self.meal_tab = ttk.Frame(nb, padding=0)
        self.tr_tab = ttk.Frame(nb, padding=0)
        nb.add(self.trip_tab, text='  行程记录  ')
        nb.add(self.lod_tab, text='  住宿记录  ')
        nb.add(self.meal_tab, text='  餐饮  ')
        nb.add(self.tr_tab, text='  公共交通  ')
        self.notebook = nb

        self._build_toolbar()
        self._build_tree()
        self._build_lodging()
        self._build_meal_transport()

    def _build_toolbar(self):
        bar = ttk.Frame(self.trip_tab, padding=(10, 8, 10, 4))
        bar.pack(fill='x')

        ttk.Button(bar, text='＋ 新增行程', command=self._add).pack(side='left')
        ttk.Button(bar, text='导出 Excel', command=self._export).pack(side='left', padx=(6, 0))
        ttk.Button(bar, text='删除所选', command=self._delete_selected).pack(side='left', padx=(6, 0))

        ttk.Separator(bar, orient='vertical').pack(side='left', fill='y', padx=12)

        ttk.Label(bar, text='从').pack(side='left')
        ttk.Entry(bar, textvariable=self.date_from_var, width=12).pack(side='left', padx=(3, 0))
        ttk.Label(bar, text='至').pack(side='left', padx=(6, 0))
        ttk.Entry(bar, textvariable=self.date_to_var, width=12).pack(side='left', padx=(3, 0))
        ttk.Button(bar, text='今天', width=6, command=self._set_today).pack(side='left', padx=(6, 0))
        ttk.Button(bar, text='本月', width=6, command=self._set_month).pack(side='left', padx=(3, 0))
        ttk.Button(bar, text='查询', width=6, command=self.refresh).pack(side='left', padx=(6, 0))
        ttk.Button(bar, text='清除', width=6, command=self._clear_filter).pack(side='left', padx=(3, 0))

        ttk.Separator(bar, orient='vertical').pack(side='left', fill='y', padx=12)

        ttk.Label(bar, text='搜索').pack(side='left')
        ent = ttk.Entry(bar, textvariable=self.keyword_var, width=18)
        ent.pack(side='left', padx=(3, 0))
        ent.bind('<Return>', lambda e: self.refresh())

        ttk.Separator(bar, orient='vertical').pack(side='left', fill='y', padx=12)
        ttk.Label(bar, text='标签').pack(side='left')
        self.trip_tag_cbo = ttk.Combobox(bar, textvariable=self.trip_tag_var, values=get_all_tags(), width=36)
        self.trip_tag_cbo.pack(side='left', padx=(3, 0))
        self.trip_tag_cbo.bind('<<ComboboxSelected>>', lambda e: self.refresh())

    # 表头顺序与固定列宽(与数据行对齐)
    HEAD_COLS = [
        ('date', '日期', 100), ('weekday', '星期', 70), ('depart', '出发地', 150),
        ('arrive', '到达地', 150), ('transport', '交通方式', 120),
        ('cost', '金额(元)', 100), ('invoice', '发票', 80),
        ('itinerary', '行程单', 80),
        ('tag', '标签', 150),
    ]
    DAY_BG = '#EAF1F8'
    DAY_BG_OPEN = '#D7E7F5'
    TOTAL_BG = '#FCE4D6'
    SEL_BG = '#BDD7EE'

    def _build_tree(self):
        wrap = ttk.Frame(self.trip_tab, padding=(10, 4, 10, 4))
        wrap.pack(fill='both', expand=True)

        self.canvas = tk.Canvas(wrap, highlightthickness=0)
        vsb = ttk.Scrollbar(wrap, orient='vertical', command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=vsb.set)
        self.canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='left', fill='y')

        inner = ttk.Frame(self.canvas, padding=0)
        self.inner = inner
        self._win = self.canvas.create_window((0, 0), window=inner, anchor='nw')

        # 表格总宽度: 所有数据列宽之和(+行内边距)
        self.table_w = sum(w for _, _, w in self.HEAD_COLS) + 16

        # 列宽: 以原始列宽为权重, 在容器变宽时按比例分摊, 撑满整个框架
        for i, (_, _, w) in enumerate(self.HEAD_COLS):
            inner.grid_columnconfigure(i, weight=w)

        # 列头
        head_bg = '#D9E1F2'
        for i, (_, text, w) in enumerate(self.HEAD_COLS):
            lbl = tk.Label(inner, text=text, width=0, anchor='center', font=FONT_BOLD,
                           bg=head_bg, relief='ridge', bd=1)
            lbl.grid(row=0, column=i, sticky='nsew')

        inner.bind('<Configure>', self._on_inner_configure)
        self.canvas.bind('<Configure>', self._on_canvas_configure)
        # 鼠标滚轮滚动(仅列表区域,不影响编辑弹窗)
        self.canvas.bind('<MouseWheel>', self._on_wheel)
        inner.bind('<MouseWheel>', self._on_wheel)

        self.day_open = {}
        self.cur_trip_id = None

    # ---------- 住宿记录区 ----------
    LOD_HEAD_COLS = [
        ('checkin', '入住日期', 120), ('checkout', '退房日期', 120),
        ('hotel', '酒店名称', 260), ('amount', '金额(元)', 100),
        ('invoice', '发票', 80), ('receipt', '水单', 80),
        ('tag', '标签', 150),
    ]
    LOD_SEL_BG = '#BDD7EE'

    def _build_lodging(self):
        wrap = ttk.Frame(self.lod_tab, padding=(10, 8, 10, 4))
        wrap.pack(fill='both', expand=True)

        bar = ttk.Frame(wrap)
        bar.pack(fill='x')
        self.lod_from_var = tk.StringVar()
        self.lod_to_var = tk.StringVar()
        self.lod_kw_var = tk.StringVar()

        ttk.Button(bar, text='＋ 新增住宿', command=self._add_lodging).pack(side='left')
        ttk.Button(bar, text='导出 Excel', command=self._export_lodging).pack(side='left', padx=(6, 0))
        ttk.Button(bar, text='删除所选', command=self._delete_lodging_selected).pack(side='left', padx=(6, 0))

        ttk.Separator(bar, orient='vertical').pack(side='left', fill='y', padx=12)
        ttk.Label(bar, text='从').pack(side='left')
        ttk.Entry(bar, textvariable=self.lod_from_var, width=12).pack(side='left', padx=(3, 0))
        ttk.Label(bar, text='至').pack(side='left', padx=(6, 0))
        ttk.Entry(bar, textvariable=self.lod_to_var, width=12).pack(side='left', padx=(3, 0))
        ttk.Button(bar, text='今天', width=5, command=self._set_lod_today).pack(side='left', padx=(6, 0))
        ttk.Button(bar, text='查询', width=5, command=self.refresh_lodging).pack(side='left', padx=(3, 0))
        ttk.Button(bar, text='清除', width=5, command=self._clear_lodging_filter).pack(side='left', padx=(3, 0))

        ttk.Separator(bar, orient='vertical').pack(side='left', fill='y', padx=12)
        ttk.Label(bar, text='酒店').pack(side='left')
        ent = ttk.Entry(bar, textvariable=self.lod_kw_var, width=14)
        ent.pack(side='left', padx=(3, 0))
        ent.bind('<Return>', lambda e: self.refresh_lodging())

        ttk.Separator(bar, orient='vertical').pack(side='left', fill='y', padx=12)
        ttk.Label(bar, text='标签').pack(side='left')
        self.lod_tag_cbo = ttk.Combobox(bar, textvariable=self.lod_tag_var, values=get_all_tags(), width=36)
        self.lod_tag_cbo.pack(side='left', padx=(3, 0))
        self.lod_tag_cbo.bind('<<ComboboxSelected>>', lambda e: self.refresh_lodging())

        body = ttk.Frame(wrap, padding=(0, 6, 0, 0))
        body.pack(fill='both', expand=True)

        self.lod_canvas = tk.Canvas(body, highlightthickness=0)
        lod_vsb = ttk.Scrollbar(body, orient='vertical', command=self.lod_canvas.yview)
        self.lod_canvas.configure(yscrollcommand=lod_vsb.set)
        self.lod_canvas.pack(side='left', fill='both', expand=True)
        lod_vsb.pack(side='left', fill='y')

        inner = ttk.Frame(self.lod_canvas, padding=0)
        self.lod_inner = inner
        self._lod_win = self.lod_canvas.create_window((0, 0), window=inner, anchor='nw')

        self.lod_table_w = sum(w for _, _, w in self.LOD_HEAD_COLS) + 8
        for i, (_, _, w) in enumerate(self.LOD_HEAD_COLS):
            inner.grid_columnconfigure(i, weight=w)

        head_bg = '#D9E1F2'
        for i, (_, text, _) in enumerate(self.LOD_HEAD_COLS):
            lbl = tk.Label(inner, text=text, anchor='center', font=FONT_BOLD,
                           bg=head_bg, relief='ridge', bd=1, height=1)
            lbl.grid(row=0, column=i, sticky='nsew')

        inner.bind('<Configure>', self._on_lod_inner_configure)
        self.lod_canvas.bind('<Configure>', self._on_lod_canvas_configure)
        self.lod_canvas.bind('<MouseWheel>', self._on_lod_wheel)
        inner.bind('<MouseWheel>', self._on_lod_wheel)

        self.cur_lod_id = None

    def _on_lod_inner_configure(self, event):
        self.lod_canvas.configure(scrollregion=self.lod_canvas.bbox('all'))

    def _on_lod_canvas_configure(self, event):
        self.lod_canvas.itemconfigure(self._lod_win, width=max(event.width, self.lod_table_w))

    def _on_lod_wheel(self, event):
        self.lod_canvas.yview_scroll(int(-event.delta / 120), 'units')

    def refresh_lodging(self):
        self._reload_tag_combos()
        for name in ('lod_from_var', 'lod_to_var'):
            v = getattr(self, name).get().strip()
            if v and parse_date(v) is None:
                messagebox.showwarning('提示', '住宿日期筛选格式应为 YYYY-MM-DD', parent=self)
                return
        raw_rows = query_lodging(self.lod_from_var.get().strip(),
                             self.lod_to_var.get().strip(),
                             self.lod_kw_var.get().strip(),
                             self.lod_tag_var.get().strip())
        # sqlite3.Row 只读,转为 dict
        rows = [dict(r) for r in raw_rows]
        # 批量获取附件计数
        lod_ids = [r['id'] for r in rows]
        file_counts = get_lodging_file_counts(lod_ids)
        for r in rows:
            fc = file_counts.get(r['id'], {'invoice': 0, 'receipt': 0})
            r['inv_cnt'] = fc['invoice']
            r['rec_cnt'] = fc['receipt']
        inner = self.lod_inner
        for child in inner.grid_slaves():
            try:
                r = int(child.grid_info()['row'])
            except Exception:
                r = -1
            if r > 0:
                child.destroy()

        row = 1
        total = 0.0
        for rec in rows:
            total += rec['amount']
            self._insert_lodging_row(inner, row, rec)
            row += 1

        if rows:
            ttl = tk.Label(inner, text='合计   %d 条住宿     总金额 ¥%.2f' % (len(rows), total),
                           anchor='center', font=FONT_BOLD, bg=self.TOTAL_BG, height=1)
            ttl.grid(row=row, column=0, columnspan=len(self.LOD_HEAD_COLS), sticky='nsew', pady=(2, 0))
            row += 1

        self._lod_summary = '住宿共 %d 条   总金额 ¥%.2f' % (len(rows), total)
        self._refresh_status()
        self.lod_canvas.configure(scrollregion=self.lod_canvas.bbox('all'))
        self.lod_canvas.yview_moveto(0)

    def _reload_tag_combos(self):
        """重新加载所有标签下拉菜单的候选值"""
        tags = get_all_tags()
        for cbo in (self.trip_tag_cbo, self.lod_tag_cbo,
                    self.meal_tag_cbo, self.tr_tag_cbo, self.calc_tag_combo):
            cbo['values'] = tags

    def _refresh_status(self):
        parts = []
        for attr in ('_trip_summary', '_lod_summary', '_meal_summary', '_transport_summary'):
            v = getattr(self, attr, '')
            if v:
                parts.append(v)
        self.status_var.set('    |    '.join(parts))

    # ---------- 餐饮 交通 区 ----------
    MEAL_HEAD_COLS = [
        ('date', '日期', 150), ('type', '餐次', 120), ('shot', '用餐截图', 200),
        ('tag', '标签', 150),
    ]
    TR_HEAD_COLS = [
        ('date', '日期', 120),
        ('depart', '出发地', 180), ('arrive', '到达地', 180),
        ('amount', '金额(元)', 120), ('shot', '交通截图', 140),
        ('tag', '标签', 150),
    ]

    def _build_meal_transport(self):
        self._build_meal_panel(self.meal_tab)
        self._build_transport_panel(self.tr_tab)

    # ---- 餐饮子页 ----
    def _build_meal_panel(self, tab):
        bar = ttk.Frame(tab); bar.pack(fill='x', padx=0, pady=(2, 4))
        ttk.Button(bar, text='＋ 新增餐饮', command=self._add_meal).pack(side='left')
        ttk.Button(bar, text='导出 Excel', command=self._export_meals).pack(side='left', padx=(6, 0))
        ttk.Button(bar, text='删除所选', command=self._delete_meal_selected).pack(side='left', padx=(6, 0))

        ttk.Separator(bar, orient='vertical').pack(side='left', fill='y', padx=12)

        ttk.Label(bar, text='从').pack(side='left')
        ttk.Entry(bar, textvariable=self.meal_date_from_var, width=12).pack(side='left', padx=(3, 0))
        ttk.Label(bar, text='至').pack(side='left', padx=(6, 0))
        ttk.Entry(bar, textvariable=self.meal_date_to_var, width=12).pack(side='left', padx=(3, 0))
        ttk.Button(bar, text='本月', width=6, command=self._set_meal_month).pack(side='left', padx=(6, 0))
        ttk.Button(bar, text='查询', width=6, command=self.refresh_meals).pack(side='left', padx=(6, 0))
        ttk.Button(bar, text='清除', width=6, command=self._clear_meal_filter).pack(side='left', padx=(3, 0))

        ttk.Separator(bar, orient='vertical').pack(side='left', fill='y', padx=12)
        ttk.Label(bar, text='标签').pack(side='left')
        self.meal_tag_cbo = ttk.Combobox(bar, textvariable=self.meal_tag_var, values=get_all_tags(), width=36)
        self.meal_tag_cbo.pack(side='left', padx=(3, 0))
        self.meal_tag_cbo.bind('<<ComboboxSelected>>', lambda e: self.refresh_meals())

        body = ttk.Frame(tab, padding=(0, 4, 0, 0))
        body.pack(fill='both', expand=True)
        self.meal_canvas = tk.Canvas(body, highlightthickness=0)
        vsb = ttk.Scrollbar(body, orient='vertical', command=self.meal_canvas.yview)
        self.meal_canvas.configure(yscrollcommand=vsb.set)
        self.meal_canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='left', fill='y')
        inner = ttk.Frame(self.meal_canvas, padding=0)
        self.meal_inner = inner
        self._meal_win = self.meal_canvas.create_window((0, 0), window=inner, anchor='nw')
        self.meal_table_w = sum(w for _, _, w in self.MEAL_HEAD_COLS) + 8
        for i, (_, _, w) in enumerate(self.MEAL_HEAD_COLS):
            inner.grid_columnconfigure(i, weight=w)
        for i, (_, text, _) in enumerate(self.MEAL_HEAD_COLS):
            tk.Label(inner, text=text, anchor='center', font=FONT_BOLD,
                     bg='#D9E1F2', relief='ridge', bd=1, height=1).grid(row=0, column=i, sticky='nsew')
        inner.bind('<Configure>', self._on_meal_inner_configure)
        self.meal_canvas.bind('<Configure>', self._on_meal_canvas_configure)
        self.meal_canvas.bind('<MouseWheel>', self._on_meal_wheel)
        inner.bind('<MouseWheel>', self._on_meal_wheel)
        self.cur_meal_id = None

    def _on_meal_inner_configure(self, event):
        self.meal_canvas.configure(scrollregion=self.meal_canvas.bbox('all'))

    def _on_meal_canvas_configure(self, event):
        self.meal_canvas.itemconfigure(self._meal_win, width=max(event.width, self.meal_table_w))

    def _on_meal_wheel(self, event):
        self.meal_canvas.yview_scroll(int(-event.delta / 120), 'units')

    def refresh_meals(self):
        self._reload_tag_combos()
        date_from = self.meal_date_from_var.get().strip()
        date_to = self.meal_date_to_var.get().strip()
        for var in (self.meal_date_from_var, self.meal_date_to_var):
            val = var.get().strip()
            if val and parse_date(val) is None:
                messagebox.showwarning('提示', '日期筛选格式应为 YYYY-MM-DD', parent=self)
                return
        rows = query_meals(date_from or '', date_to or '', self.meal_tag_var.get().strip())
        inner = self.meal_inner
        for child in inner.grid_slaves():
            try:
                r = int(child.grid_info()['row'])
            except Exception:
                r = -1
            if r > 0:
                child.destroy()
        row = 1
        for rec in rows:
            self._insert_meal_row(inner, row, rec); row += 1
        if rows:
            tk.Label(inner, text='共 %d 条餐饮记录' % len(rows),
                     anchor='center', font=FONT_BOLD, bg=self.TOTAL_BG, height=1).grid(
                row=row, column=0, columnspan=len(self.MEAL_HEAD_COLS), sticky='nsew', pady=(2, 0))
        self._meal_summary = '餐饮共 %d 条' % len(rows)
        self._refresh_status()
        self.meal_canvas.configure(scrollregion=self.meal_canvas.bbox('all'))
        self.meal_canvas.yview_moveto(0)

    def _insert_meal_row(self, parent, row, rec):
        values = (rec['meal_date'], rec['meal_type'] or '—',
                  '有截图' if rec['screenshot'] else '—',
                  rec['trip_tag'] or '—')
        bg = self.SEL_BG if self.cur_meal_id == rec['id'] else '#FFFFFF'
        cells = []
        for i, v in enumerate(values):
            lbl = tk.Label(parent, text=v, bg=bg, height=1, font=FONT, bd=0, anchor='center')
            lbl.grid(row=row, column=i, sticky='nsew'); cells.append(lbl)
        for c in cells:
            c.bind('<Button-1>', lambda e, mid=rec['id']: self._select_meal(mid))
            c.bind('<Double-1>', lambda e, mid=rec['id']: self._edit_meal(mid))
            c.bind('<Button-3>', lambda e, mid=rec['id']: self._on_right_click_meal(e, mid))

    def _select_meal(self, mid):
        if self.cur_meal_id == mid:
            return
        self.cur_meal_id = mid
        self.refresh_meals()

    def _add_meal(self):
        dlg = MealDialog(self)
        self.wait_window(dlg)
        self.refresh_meals()

    def _edit_meal(self, mid):
        rec = get_meal(mid)
        if rec:
            dlg = MealDialog(self, dict(rec))
            self.wait_window(dlg)
            self.refresh_meals()

    def _export_meals(self):
        rows = query_meals(self.meal_date_from_var.get().strip() or '',
                           self.meal_date_to_var.get().strip() or '',
                           self.meal_tag_var.get().strip())
        path, n = export_meals_to_excel(rows)
        if path is None:
            messagebox.showinfo('导出', '当前没有餐饮记录可导出', parent=self); return
        if messagebox.askyesno('导出成功', '已导出 %d 条餐饮记录到:\n%s\n\n是否立即打开?' % (n, path), parent=self):
            try:
                os.startfile(path)
            except OSError:
                pass

    def _delete_meal_selected(self):
        mid = self.cur_meal_id
        if not mid:
            messagebox.showinfo('提示', '请先选中要删除的餐饮记录', parent=self); return
        if messagebox.askyesno('删除确认', '确定删除这条餐饮记录吗?', parent=self):
            delete_meal(mid)
            self.cur_meal_id = None
            self.refresh_meals()

    def _on_right_click_meal(self, event, mid):
        self._select_meal(mid)
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label='编辑', command=lambda: self._edit_meal(mid))
        menu.add_command(label='删除', command=self._delete_meal_selected)
        menu.tk_popup(event.x_root, event.y_root)

    # ---- 交通子页 ----
    def _build_transport_panel(self, tab):
        bar = ttk.Frame(tab); bar.pack(fill='x', pady=(2, 4))
        ttk.Button(bar, text='＋ 新增交通', command=self._add_transport).pack(side='left')
        ttk.Button(bar, text='导出 Excel', command=self._export_transports).pack(side='left', padx=(6, 0))
        ttk.Button(bar, text='删除所选', command=self._delete_transport_selected).pack(side='left', padx=(6, 0))

        ttk.Separator(bar, orient='vertical').pack(side='left', fill='y', padx=12)

        ttk.Label(bar, text='从').pack(side='left')
        ttk.Entry(bar, textvariable=self.tr_date_from_var, width=12).pack(side='left', padx=(3, 0))
        ttk.Label(bar, text='至').pack(side='left', padx=(6, 0))
        ttk.Entry(bar, textvariable=self.tr_date_to_var, width=12).pack(side='left', padx=(3, 0))
        ttk.Button(bar, text='本月', width=6, command=self._set_tr_month).pack(side='left', padx=(6, 0))
        ttk.Button(bar, text='查询', width=6, command=self.refresh_transports).pack(side='left', padx=(6, 0))
        ttk.Button(bar, text='清除', width=6, command=self._clear_tr_filter).pack(side='left', padx=(3, 0))

        ttk.Separator(bar, orient='vertical').pack(side='left', fill='y', padx=12)
        ttk.Label(bar, text='标签').pack(side='left')
        self.tr_tag_cbo = ttk.Combobox(bar, textvariable=self.tr_tag_var, values=get_all_tags(), width=36)
        self.tr_tag_cbo.pack(side='left', padx=(3, 0))
        self.tr_tag_cbo.bind('<<ComboboxSelected>>', lambda e: self.refresh_transports())

        body = ttk.Frame(tab, padding=(0, 4, 0, 0))
        body.pack(fill='both', expand=True)
        self.tr_canvas = tk.Canvas(body, highlightthickness=0)
        vsb = ttk.Scrollbar(body, orient='vertical', command=self.tr_canvas.yview)
        self.tr_canvas.configure(yscrollcommand=vsb.set)
        self.tr_canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='left', fill='y')
        inner = ttk.Frame(self.tr_canvas, padding=0)
        self.tr_inner = inner
        self._tr_win = self.tr_canvas.create_window((0, 0), window=inner, anchor='nw')
        self.tr_table_w = sum(w for _, _, w in self.TR_HEAD_COLS) + 8
        for i, (_, _, w) in enumerate(self.TR_HEAD_COLS):
            inner.grid_columnconfigure(i, weight=w)
        for i, (_, text, _) in enumerate(self.TR_HEAD_COLS):
            tk.Label(inner, text=text, anchor='center', font=FONT_BOLD,
                     bg='#D9E1F2', relief='ridge', bd=1, height=1).grid(row=0, column=i, sticky='nsew')
        inner.bind('<Configure>', self._on_tr_inner_configure)
        self.tr_canvas.bind('<Configure>', self._on_tr_canvas_configure)
        self.tr_canvas.bind('<MouseWheel>', self._on_tr_wheel)
        inner.bind('<MouseWheel>', self._on_tr_wheel)
        self.cur_tr_id = None

    def _on_tr_inner_configure(self, event):
        self.tr_canvas.configure(scrollregion=self.tr_canvas.bbox('all'))

    def _on_tr_canvas_configure(self, event):
        self.tr_canvas.itemconfigure(self._tr_win, width=max(event.width, self.tr_table_w))

    def _on_tr_wheel(self, event):
        self.tr_canvas.yview_scroll(int(-event.delta / 120), 'units')

    def refresh_transports(self):
        self._reload_tag_combos()
        date_from = self.tr_date_from_var.get().strip()
        date_to = self.tr_date_to_var.get().strip()
        for var in (self.tr_date_from_var, self.tr_date_to_var):
            val = var.get().strip()
            if val and parse_date(val) is None:
                messagebox.showwarning('提示', '日期筛选格式应为 YYYY-MM-DD', parent=self)
                return
        rows = query_transports(date_from or '', date_to or '', self.tr_tag_var.get().strip())
        inner = self.tr_inner
        for child in inner.grid_slaves():
            try:
                r = int(child.grid_info()['row'])
            except Exception:
                r = -1
            if r > 0:
                child.destroy()
        row = 1
        total = 0.0
        for rec in rows:
            total += rec['amount']
            self._insert_transport_row(inner, row, rec); row += 1
        if rows:
            tk.Label(inner, text='合计   %d 条交通     总金额 ¥%.2f' % (len(rows), total),
                     anchor='center', font=FONT_BOLD, bg=self.TOTAL_BG, height=1).grid(
                row=row, column=0, columnspan=len(self.TR_HEAD_COLS), sticky='nsew', pady=(2, 0))
        self._transport_summary = '交通共 %d 条   总金额 ¥%.2f' % (len(rows), total)
        self._refresh_status()
        self.tr_canvas.configure(scrollregion=self.tr_canvas.bbox('all'))
        self.tr_canvas.yview_moveto(0)

    def _insert_transport_row(self, parent, row, rec):
        values = (rec['t_date'], rec['depart'] or '—', rec['arrive'] or '—',
                  '¥%.2f' % rec['amount'],
                  '有截图' if rec['screenshot'] else '—',
                  rec['trip_tag'] or '—')
        bg = self.SEL_BG if self.cur_tr_id == rec['id'] else '#FFFFFF'
        cells = []
        for i, v in enumerate(values):
            lbl = tk.Label(parent, text=v, bg=bg, height=1, font=FONT, bd=0, anchor='center')
            lbl.grid(row=row, column=i, sticky='nsew'); cells.append(lbl)
        for c in cells:
            c.bind('<Button-1>', lambda e, tid=rec['id']: self._select_transport(tid))
            c.bind('<Double-1>', lambda e, tid=rec['id']: self._edit_transport(tid))
            c.bind('<Button-3>', lambda e, tid=rec['id']: self._on_right_click_transport(e, tid))

    def _select_transport(self, tid):
        if self.cur_tr_id == tid:
            return
        self.cur_tr_id = tid
        self.refresh_transports()

    def _add_transport(self):
        dlg = TransportDialog(self)
        self.wait_window(dlg)
        self.refresh_transports()

    def _edit_transport(self, tid):
        rec = get_transport(tid)
        if rec:
            dlg = TransportDialog(self, dict(rec))
            self.wait_window(dlg)
            self.refresh_transports()

    def _export_transports(self):
        rows = query_transports(self.tr_date_from_var.get().strip() or '',
                                self.tr_date_to_var.get().strip() or '',
                                self.tr_tag_var.get().strip())
        path, n = export_transports_to_excel(rows)
        if path is None:
            messagebox.showinfo('导出', '当前没有交通记录可导出', parent=self); return
        if messagebox.askyesno('导出成功', '已导出 %d 条交通记录到:\n%s\n\n是否立即打开?' % (n, path), parent=self):
            try:
                os.startfile(path)
            except OSError:
                pass

    def _delete_transport_selected(self):
        tid = self.cur_tr_id
        if not tid:
            messagebox.showinfo('提示', '请先选中要删除的交通记录', parent=self); return
        if messagebox.askyesno('删除确认', '确定删除这条交通记录吗?', parent=self):
            delete_transport(tid)
            self.cur_tr_id = None
            self.refresh_transports()

    def _on_right_click_transport(self, event, tid):
        self._select_transport(tid)
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label='编辑', command=lambda: self._edit_transport(tid))
        menu.add_command(label='删除', command=self._delete_transport_selected)
        menu.tk_popup(event.x_root, event.y_root)

    def _insert_lodging_row(self, parent, row, rec):
        inv_cnt = rec.get('inv_cnt', 0)
        rec_cnt = rec.get('rec_cnt', 0)
        inv_text = '发票(%d)' % inv_cnt if inv_cnt else '—'
        rec_text = '水单(%d)' % rec_cnt if rec_cnt else '—'
        values = (rec['checkin_date'], rec['checkout_date'], rec['hotel'],
                  '¥%.2f' % rec['amount'], inv_text, rec_text, rec['trip_tag'] or '—')
        bg = self.LOD_SEL_BG if self.cur_lod_id == rec['id'] else '#FFFFFF'
        cells = []
        for i, v in enumerate(values):
            is_link = (i == 4 and inv_cnt) or (i == 5 and rec_cnt)
            lbl = tk.Label(parent, text=v, bg=bg, height=1, font=FONT, bd=0,
                           anchor='w' if i == 2 else 'center',
                           foreground='#0563C1' if is_link else 'black',
                           cursor='hand2' if is_link else 'arrow')
            if is_link:
                lbl.configure(font=(FONT[0], FONT[1], 'underline'))
            lbl.grid(row=row, column=i, sticky='nsew')
            cells.append(lbl)
        for i, c in enumerate(cells):
            is_link = (i == 4 and inv_cnt) or (i == 5 and rec_cnt)
            if is_link:
                ft = 'invoice' if i == 4 else 'receipt'
                c.bind('<Button-1>', lambda e, lid=rec['id'], ft=ft: self._show_lod_files_preview(lid, ft))
            else:
                c.bind('<Button-1>', lambda e, lid=rec['id']: self._select_lodging(lid))
            c.bind('<Double-1>', lambda e, lid=rec['id']: self._edit_lodging(lid))
            c.bind('<Button-3>', lambda e, lid=rec['id']: self._on_right_click_lodging(e, lid))

    def _show_lod_files_preview(self, lodging_id, file_type):
        tag = '发票' if file_type == 'invoice' else '水单'
        files = [f for f in get_lodging_files(lodging_id) if f['file_type'] == file_type]
        if not files:
            messagebox.showinfo('提示', '暂无%s文件' % tag, parent=self)
            return
        win = tk.Toplevel(self)
        win.title('%s列表' % tag)
        win.transient(self)
        ttk.Label(win, text='%s (%d 个文件):' % (tag, len(files)),
                  font=FONT_BOLD, padding=10).pack(anchor='w')
        for f in files:
            abs_path = rel_to_abs(f['file_path'])
            name = f['file_name'] or os.path.basename(f['file_path'])
            if len(name) > 40:
                name = name[:37] + '...'
            btn = ttk.Button(win, text=name, width=45,
                             command=lambda p=abs_path: self._open_file(p))
            btn.pack(padx=10, pady=2, fill='x')
        win.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() - win.winfo_width()) // 2
        y = self.winfo_rooty() + (self.winfo_height() - win.winfo_height()) // 2
        win.geometry('+%d+%d' % (max(x, 0), max(y, 0)))

    def _select_lodging(self, lid):
        if self.cur_lod_id == lid:
            return
        self.cur_lod_id = lid
        self.refresh_lodging()

    def _add_lodging(self):
        dlg = LodgingDialog(self)
        self.wait_window(dlg)
        self.refresh_lodging()

    def _edit_lodging(self, lid):
        rec = get_lodging(lid)
        if rec:
            dlg = LodgingDialog(self, dict(rec))
            self.wait_window(dlg)
            self.refresh_lodging()

    def _export_lodging(self):
        rows = query_lodging(self.lod_from_var.get().strip(),
                             self.lod_to_var.get().strip(),
                             self.lod_kw_var.get().strip(),
                             self.lod_tag_var.get().strip())
        path, n = export_lodging_to_excel(rows)
        if path is None:
            messagebox.showinfo('导出', '当前筛选范围内没有住宿记录可导出', parent=self)
            return
        ret = messagebox.askyesno('导出成功', '已导出 %d 条住宿记录到:\n%s\n\n是否立即打开?' % (n, path), parent=self)
        if ret:
            try:
                os.startfile(path)
            except OSError:
                pass

    def _set_lod_today(self):
        self.lod_from_var.set(datetime.date.today().isoformat())
        self.lod_to_var.set(datetime.date.today().isoformat())
        self.refresh_lodging()

    def _clear_lodging_filter(self):
        self.lod_from_var.set('')
        self.lod_to_var.set('')
        self.lod_kw_var.set('')
        self.lod_tag_var.set('')
        self.refresh_lodging()

    def _delete_lodging_selected(self):
        lid = self.cur_lod_id
        if not lid:
            messagebox.showinfo('提示', '请先选中要删除的住宿记录', parent=self)
            return
        rec = get_lodging(lid)
        if rec is None:
            return
        ans = messagebox.askyesno('删除确认', '确定删除这条住宿记录吗?', parent=self)
        if not ans:
            return
        delete_lodging(lid)
        self.cur_lod_id = None
        self.refresh_lodging()

    def _on_right_click_lodging(self, event, lid):
        self._select_lodging(lid)
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label='编辑', command=lambda: self._edit_lodging(lid))
        menu.add_command(label='删除', command=self._delete_lodging_selected)
        menu.tk_popup(event.x_root, event.y_root)

    def _on_inner_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox('all'))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self._win, width=max(event.width, self.table_w))

    def _on_wheel(self, event):
        self.canvas.yview_scroll(int(-event.delta / 120), 'units')

    def _build_calc_panel(self):
        """计算出差费用总额面板, 位于 notebook 与状态栏之间"""
        ttk.Separator(self).pack(fill='x')
        frame = ttk.Frame(self, padding=(10, 6))
        frame.pack(fill='x')

        ttk.Label(frame, text='标签:').pack(side='left')
        self.calc_tag_var = tk.StringVar()
        self.calc_tag_combo = ttk.Combobox(frame, textvariable=self.calc_tag_var, width=30)
        self.calc_tag_combo.pack(side='left', padx=(3, 0))
        self.calc_tag_combo.bind('<FocusIn>', lambda e: self._refresh_calc_tags())

        ttk.Button(frame, text='计算出差费用总额', command=self._do_calc).pack(side='left', padx=(8, 0))

        self.calc_result_var = tk.StringVar()
        self.calc_result_entry = ttk.Entry(frame, textvariable=self.calc_result_var, width=65, state='readonly')
        self.calc_result_entry.pack(side='left', padx=(8, 0), fill='x', expand=True)

    def _refresh_calc_tags(self):
        tags = get_all_tags()
        self.calc_tag_combo['values'] = tags

    def _do_calc(self):
        tag = self.calc_tag_var.get().strip()
        if not tag:
            messagebox.showwarning('提示', '请先选择出差标签')
            return
        try:
            r = calc_trip_total(tag)
        except Exception as e:
            messagebox.showerror('计算错误', str(e))
            return

        lines = [
            '%s (%s城市: %s)' % (tag, r['tier_name'], r['city'] or '未知'),
            '行程: %d笔 ¥%.2f' % (r['trip_count'], r['trip_total']),
            '住宿: %d笔 ¥%.2f' % (r['lod_count'], r['lod_total']),
            '餐费: %d天×¥%d=¥%.2f' % (r['meal_days'], r['meal_rate'], r['meal_cost']),
            '市内交通: %d天×¥%d=¥%.2f' % (r['transit_days'], r['transit_rate'], r['transit_cost']),
            '合计: ¥%.2f' % r['total'],
        ]
        self.calc_result_var.set(' | '.join(lines))

    def _build_statusbar(self):
        self.status_var = tk.StringVar()
        ttk.Separator(self).pack(fill='x')
        ttk.Label(self, textvariable=self.status_var, anchor='w',
                  padding=(12, 4)).pack(fill='x')

    # ---------- 数据刷新 ----------
    def refresh(self):
        self._reload_tag_combos()
        date_from = self.date_from_var.get().strip()
        date_to = self.date_to_var.get().strip()
        for name in ('date_from_var', 'date_to_var'):
            if getattr(self, name).get().strip() and parse_date(getattr(self, name).get()) is None:
                messagebox.showwarning('提示', '日期筛选格式应为 YYYY-MM-DD', parent=self)
                return
        rows = query_trips(date_from or '', date_to or '', self.keyword_var.get(),
                           self.trip_tag_var.get().strip())
        # sqlite3.Row 只读,转为 dict 以便附加附件计数
        rows = [dict(r) for r in rows]

        # 批量获取附件计数
        trip_ids = [r['id'] for r in rows]
        file_counts = get_trip_file_counts(trip_ids)
        for r in rows:
            fc = file_counts.get(r['id'], {'invoice': 0, 'itinerary': 0})
            r['inv_cnt'] = fc['invoice']
            r['it_cnt'] = fc['itinerary']

        inner = self.inner
        # 清空列头以下的所有行(保留第 0 行列头)
        for child in inner.grid_slaves():
            try:
                r = int(child.grid_info()['row'])
            except Exception:
                r = -1
            if r > 0:
                child.destroy()

        day_map = {}
        for r in rows:
            d = r['trip_date']
            day_map.setdefault(d, []).append(r)

        row = 1
        for day, day_rows in day_map.items():
            opened = self.day_open.get(day, True)
            # —— 日分组合并行(跨整行) ——
            tri = '▾ %s' % day if opened else '▸ %s' % day
            text = '%s   %s    |   %d 条行程' % (tri, weekday_of(day), len(day_rows))
            day_lbl = tk.Label(inner, text=text, anchor='w', font=FONT_BOLD,
                               bg=self.DAY_BG_OPEN if opened else self.DAY_BG,
                               padx=8, height=1)
            day_lbl.grid(row=row, column=0, columnspan=len(self.HEAD_COLS),
                         sticky='nsew', pady=(2, 0))
            day_lbl.configure(cursor='hand2')
            day_lbl.bind('<Button-1>', lambda e, d_=day: self._toggle_day(d_))
            row += 1

            # 组内行程行
            self.day_open[day] = opened
            if opened:
                for r in day_rows:
                    self._insert_trip_row(inner, row, r)
                    row += 1

        # —— 总条数合并行(跨整行) ——
        if day_map:
            total_cost = sum(r['cost'] for r in rows)
            ttl_text = '合计   %d 条行程' % len(rows)
            if total_cost:
                ttl_text += '     总金额 ¥%.2f' % total_cost
            ttl = tk.Label(inner,
                           text=ttl_text,
                           anchor='center', font=FONT_BOLD, bg=self.TOTAL_BG, height=1)
            ttl.grid(row=row, column=0, columnspan=len(self.HEAD_COLS),
                     sticky='nsew', pady=(2, 0))
            row += 1

        flt = '全部记录' if not date_from and not date_to else ('%s ~ %s' % (date_from or '最早', date_to or '至今'))
        total_cost = sum(r['cost'] for r in rows)
        self._trip_summary = '共 %d 条行程   筛选范围: %s' % (len(rows), flt)
        if total_cost:
            self._trip_summary += '   总金额 ¥%.2f' % total_cost
        self._refresh_status()
        self.canvas.configure(scrollregion=self.canvas.bbox('all'))
        self.canvas.yview_moveto(0)

    def _insert_trip_row(self, parent, row, rec):
        cost_val = rec['cost']
        inv_cnt = rec.get('inv_cnt', 0)
        it_cnt = rec.get('it_cnt', 0)
        inv_text = '发票(%d)' % inv_cnt if inv_cnt else '—'
        it_text = '行程单(%d)' % it_cnt if it_cnt else '—'
        values = (rec['trip_date'], rec['weekday'], rec['depart'], rec['arrive'],
                  rec['transport'],
                  '¥%.2f' % cost_val if cost_val else '—',
                  inv_text, it_text, rec['trip_tag'] or '—')
        bg = self.SEL_BG if self.cur_trip_id == rec['id'] else '#FFFFFF'
        cells = []
        for i, v in enumerate(values):
            is_link = (i == 6 and inv_cnt) or (i == 7 and it_cnt)
            lbl = tk.Label(parent, text=v, anchor='center', bg=bg, height=1,
                           font=FONT, bd=0,
                           foreground='#0563C1' if is_link else 'black',
                           cursor='hand2' if is_link else 'arrow')
            if is_link:
                lbl.configure(font=(FONT[0], FONT[1], 'underline'))
            lbl.grid(row=row, column=i, sticky='nsew')
            cells.append(lbl)
        for i, c in enumerate(cells):
            is_link = (i == 6 and inv_cnt) or (i == 7 and it_cnt)
            if is_link:
                ft = 'invoice' if i == 6 else 'itinerary'
                c.bind('<Button-1>', lambda e, tid=rec['id'], ft=ft: self._show_files_preview(tid, ft))
            else:
                c.bind('<Button-1>', lambda e, tid=rec['id']: self._select_trip(tid))
            c.bind('<Double-1>', lambda e, tid=rec['id']: self._edit_by_id(tid))
            c.bind('<Button-3>', lambda e, tid=rec['id']: self._on_right_click_row(e, tid))

    def _show_files_preview(self, trip_id, file_type):
        """弹出附件列表窗口，每个文件可点击打开"""
        tag = '发票' if file_type == 'invoice' else '电子行程单'
        files = [f for f in get_trip_files(trip_id) if f['file_type'] == file_type]
        if not files:
            messagebox.showinfo('提示', '暂无%s文件' % tag, parent=self)
            return
        win = tk.Toplevel(self)
        win.title('%s列表' % tag)
        win.transient(self)
        ttk.Label(win, text='%s (%d 个文件):' % (tag, len(files)),
                  font=FONT_BOLD, padding=10).pack(anchor='w')
        for f in files:
            abs_path = rel_to_abs(f['file_path'])
            name = f['file_name'] or os.path.basename(f['file_path'])
            if len(name) > 40:
                name = name[:37] + '...'
            btn = ttk.Button(win, text=name, width=45,
                             command=lambda p=abs_path: self._open_file(p))
            btn.pack(padx=10, pady=2, fill='x')
        win.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() - win.winfo_width()) // 2
        y = self.winfo_rooty() + (self.winfo_height() - win.winfo_height()) // 2
        win.geometry('+%d+%d' % (max(x, 0), max(y, 0)))

    def _open_file(self, abs_path):
        """打开单个文件: PDF用系统默认程序，图片弹窗预览"""
        if not abs_path or not os.path.isfile(abs_path):
            messagebox.showinfo('提示', '文件不存在', parent=self)
            return
        ext = os.path.splitext(abs_path)[1].lower()
        if ext == '.pdf':
            try:
                os.startfile(abs_path)
            except OSError as e:
                messagebox.showerror('提示', '无法打开 PDF: %s' % str(e), parent=self)
            return
        win = tk.Toplevel(self)
        win.title('发票预览')
        win.transient(self)
        try:
            img = Image.open(abs_path)
            max_w, max_h = 800, 900
            w, h = img.size
            if w > max_w or h > max_h:
                ratio = min(max_w / w, max_h / h)
                img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            canvas = tk.Canvas(win, width=photo.width(), height=photo.height(),
                               highlightthickness=0)
            canvas.pack(padx=10, pady=10)
            canvas.create_image(0, 0, anchor='nw', image=photo)
            canvas.image = photo
        except Exception as e:
            ttk.Label(win, text='无法加载图片: %s' % str(e), padding=20).pack()
        win.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() - win.winfo_width()) // 2
        y = self.winfo_rooty() + (self.winfo_height() - win.winfo_height()) // 2
        win.geometry('+%d+%d' % (max(x, 0), max(y, 0)))

    def _select_trip(self, trip_id):
        if self.cur_trip_id == trip_id:
            return
        self.cur_trip_id = trip_id
        self.refresh()

    def _toggle_day(self, day):
        self.day_open[day] = not self.day_open.get(day, True)
        self.refresh()

    def _edit_by_id(self, trip_id):
        rec = get_trip(trip_id)
        if rec:
            dlg = EditDialog(self, dict(rec))
            self.wait_window(dlg)
            self.refresh()

    def _on_right_click_row(self, event, trip_id):
        self._select_trip(trip_id)
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label='编辑', command=lambda: self._edit_by_id(trip_id))
        menu.add_command(label='删除', command=self._delete_selected)
        menu.tk_popup(event.x_root, event.y_root)

    # ---------- 工具栏行为 ----------
    def _add(self):
        dlg = EditDialog(self)
        self.wait_window(dlg)
        self.refresh()

    def _set_today(self):
        today = datetime.date.today().isoformat()
        self.date_from_var.set(today)
        self.date_to_var.set(today)
        self.refresh()

    def _set_month(self):
        today = datetime.date.today()
        self.date_from_var.set(today.replace(day=1).isoformat())
        self.date_to_var.set(today.isoformat())
        self.refresh()

    def _clear_filter(self):
        self.date_from_var.set('')
        self.date_to_var.set('')
        self.keyword_var.set('')
        self.trip_tag_var.set('')
        self.refresh()

    # ---- 餐饮/交通筛选辅助方法 ----
    def _set_meal_month(self):
        today = datetime.date.today()
        self.meal_date_from_var.set(today.replace(day=1).isoformat())
        self.meal_date_to_var.set(today.isoformat())
        self.refresh_meals()

    def _clear_meal_filter(self):
        self.meal_date_from_var.set('')
        self.meal_date_to_var.set('')
        self.meal_tag_var.set('')
        self.refresh_meals()

    def _set_tr_month(self):
        today = datetime.date.today()
        self.tr_date_from_var.set(today.replace(day=1).isoformat())
        self.tr_date_to_var.set(today.isoformat())
        self.refresh_transports()

    def _clear_tr_filter(self):
        self.tr_date_from_var.set('')
        self.tr_date_to_var.set('')
        self.tr_tag_var.set('')
        self.refresh_transports()

    def _export(self):
        rows = query_trips(self.date_from_var.get().strip() or '',
                           self.date_to_var.get().strip() or '',
                           self.keyword_var.get(),
                           self.trip_tag_var.get().strip())
        path, n = export_to_excel(rows)
        if path is None:
            messagebox.showinfo('导出', '当前筛选范围内没有可导出的记录', parent=self)
            return
        ret = messagebox.askyesno('导出成功', '已导出 %d 条记录到:\n%s\n\n是否立即打开?' % (n, path), parent=self)
        if ret:
            try:
                os.startfile(path)
            except OSError:
                pass

    def _delete_selected(self):
        trip_id = self.cur_trip_id
        if not trip_id:
            messagebox.showinfo('提示', '请先选中要删除的具体记录', parent=self)
            return
        rec = get_trip(trip_id)
        if rec is None:
            return
        ans = messagebox.askyesnocancel(
            '删除确认',
            '确定删除这条记录吗?\n\n'
            '【是】删除记录并同时删除关联截图文件\n'
            '【否】仅删除记录,保留截图文件\n'
            '【取消】不删除',
            parent=self)
        if ans is None:
            return
        delete_trip(rec['id'], bool(ans))
        self.cur_trip_id = None
        self.refresh()


# ---------------- 入口 ----------------

def main():
    try:
        db_init()
        app = MainApp()
        app.mainloop()
    except Exception as e:
        try:
            messagebox.showerror('程序错误', str(e))
        except Exception:
            raise


if __name__ == '__main__':
    main()